from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Prefetch
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

from .models import JobCategory, JobType, JobRole, JobProfile, JobProfileHistory
from .models import UserJobProfileBookmark, UserJobProfileView
from .forms import JobProfileForm


def job_profile_list(request):
    """직무기술서 목록 (ESS용 - 개선된 버전)"""
    # 검색 필터
    search_query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    job_type_id = request.GET.get('job_type', '')
    
    # 정렬 옵션 (신규)
    sort_by = request.GET.get('sort', 'name')  # name, created_at, updated_at
    order = request.GET.get('order', 'asc')  # asc, desc
    
    # 기본 쿼리셋
    profiles = JobProfile.objects.filter(is_active=True).select_related(
        'job_role__job_type__category'
    ).prefetch_related('job_role')
    
    # 검색어 필터
    if search_query:
        profiles = profiles.filter(
            Q(job_role__name__icontains=search_query) |
            Q(role_responsibility__icontains=search_query) |
            Q(qualification__icontains=search_query)
        )
    
    # 직군 필터
    if category_id:
        profiles = profiles.filter(job_role__job_type__category_id=category_id)
    
    # 직종 필터
    if job_type_id:
        profiles = profiles.filter(job_role__job_type_id=job_type_id)
    
    # 검색 결과 수 (신규)
    total_count = profiles.count()
    
    # 정렬 적용 (신규)
    order_prefix = '-' if order == 'desc' else ''
    if sort_by == 'name':
        profiles = profiles.order_by(f'{order_prefix}job_role__name')
    elif sort_by == 'created_at':
        profiles = profiles.order_by(f'{order_prefix}created_at')
    elif sort_by == 'updated_at':
        profiles = profiles.order_by(f'{order_prefix}updated_at')
    
    # 사용자 북마크 정보 (신규)
    bookmarked_ids = []
    if request.user.is_authenticated:
        bookmarked_ids = list(
            UserJobProfileBookmark.objects.filter(
                user=request.user,
                is_active=True
            ).values_list('job_profile_id', flat=True)
        )
    
    # 페이지네이션
    paginator = Paginator(profiles, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 필터용 데이터
    categories = JobCategory.objects.filter(is_active=True)
    job_types = []
    if category_id:
        job_types = JobType.objects.filter(category_id=category_id, is_active=True)
    
    # 최근 본 직무 (신규)
    recent_views = []
    if request.user.is_authenticated:
        recent_views = UserJobProfileView.objects.filter(
            user=request.user
        ).select_related(
            'job_profile__job_role__job_type__category'
        ).order_by('-viewed_at')[:5]
    
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'job_types': job_types,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_job_type': job_type_id,
        'sort_by': sort_by,
        'order': order,
        'total_count': total_count,  # 신규
        'bookmarked_ids': bookmarked_ids,  # 신규
        'recent_views': recent_views,  # 신규
    }
    
    return render(request, 'job_profiles/job_profile_list.html', context)


@login_required
def job_profile_download(request):
    """직무기술서 목록 다운로드 (신규)"""
    # 필터 파라미터 (목록과 동일)
    search_query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    job_type_id = request.GET.get('job_type', '')
    format_type = request.GET.get('format', 'excel')  # excel, csv
    
    # 쿼리셋 구성
    profiles = JobProfile.objects.filter(is_active=True).select_related(
        'job_role__job_type__category'
    )
    
    if search_query:
        profiles = profiles.filter(
            Q(job_role__name__icontains=search_query) |
            Q(role_responsibility__icontains=search_query) |
            Q(qualification__icontains=search_query)
        )
    
    if category_id:
        profiles = profiles.filter(job_role__job_type__category_id=category_id)
    
    if job_type_id:
        profiles = profiles.filter(job_role__job_type_id=job_type_id)
    
    # 다운로드 형식에 따른 처리
    if format_type == 'csv':
        return download_csv(profiles)
    else:
        return download_excel(profiles)


def download_csv(profiles):
    """CSV 다운로드"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="job_profiles_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        '직군', '직종', '직무', '직무코드',
        '역할과 책임', '자격요건', '기본 스킬',
        '응용 스킬', '성장경로', '관련 자격증',
        '생성일', '수정일'
    ])
    
    for profile in profiles:
        writer.writerow([
            profile.job_role.job_type.category.name,
            profile.job_role.job_type.name,
            profile.job_role.name,
            profile.job_role.code,
            profile.role_responsibility,
            profile.qualification,
            ', '.join(profile.basic_skills),
            ', '.join(profile.applied_skills),
            profile.growth_path,
            ', '.join(profile.related_certifications),
            profile.created_at.strftime('%Y-%m-%d'),
            profile.updated_at.strftime('%Y-%m-%d')
        ])
    
    return response


def download_excel(profiles):
    """Excel 다운로드"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "직무기술서"
    
    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = [
        '직군', '직종', '직무', '직무코드',
        '역할과 책임', '자격요건', '기본 스킬',
        '응용 스킬', '성장경로', '관련 자격증',
        '생성일', '수정일'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 데이터 작성
    for row, profile in enumerate(profiles, 2):
        ws.cell(row=row, column=1, value=profile.job_role.job_type.category.name)
        ws.cell(row=row, column=2, value=profile.job_role.job_type.name)
        ws.cell(row=row, column=3, value=profile.job_role.name)
        ws.cell(row=row, column=4, value=profile.job_role.code)
        ws.cell(row=row, column=5, value=profile.role_responsibility)
        ws.cell(row=row, column=6, value=profile.qualification)
        ws.cell(row=row, column=7, value=', '.join(profile.basic_skills))
        ws.cell(row=row, column=8, value=', '.join(profile.applied_skills))
        ws.cell(row=row, column=9, value=profile.growth_path)
        ws.cell(row=row, column=10, value=', '.join(profile.related_certifications))
        ws.cell(row=row, column=11, value=profile.created_at.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=12, value=profile.updated_at.strftime('%Y-%m-%d'))
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # BytesIO에 저장
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="job_profiles_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


@login_required
@csrf_exempt
def toggle_bookmark(request, profile_id):
    """북마크 토글 (신규)"""
    if request.method == 'POST':
        profile = get_object_or_404(JobProfile, id=profile_id)
        
        bookmark, created = UserJobProfileBookmark.objects.get_or_create(
            user=request.user,
            job_profile=profile,
            defaults={'is_active': True}
        )
        
        if not created:
            bookmark.is_active = not bookmark.is_active
            bookmark.save()
        
        return JsonResponse({
            'success': True,
            'bookmarked': bookmark.is_active,
            'message': '북마크에 추가되었습니다.' if bookmark.is_active else '북마크가 해제되었습니다.'
        })
    
    return JsonResponse({'success': False, 'message': '잘못된 요청입니다.'})


@login_required
def job_profile_detail(request, profile_id):
    """직무기술서 상세보기 (개선)"""
    profile = get_object_or_404(
        JobProfile.objects.select_related('job_role__job_type__category'),
        id=profile_id,
        is_active=True
    )
    
    # 조회 기록 저장 (신규)
    if request.user.is_authenticated:
        UserJobProfileView.objects.update_or_create(
            user=request.user,
            job_profile=profile,
            defaults={'viewed_at': timezone.now()}
        )
    
    # 북마크 여부 확인 (신규)
    is_bookmarked = False
    if request.user.is_authenticated:
        is_bookmarked = UserJobProfileBookmark.objects.filter(
            user=request.user,
            job_profile=profile,
            is_active=True
        ).exists()
    
    # 관련 직무들
    related_profiles = JobProfile.objects.filter(
        job_role__job_type=profile.job_role.job_type,
        is_active=True
    ).exclude(id=profile.id).select_related('job_role')[:5]
    
    context = {
        'profile': profile,
        'related_profiles': related_profiles,
        'is_bookmarked': is_bookmarked,  # 신규
    }
    
    return render(request, 'job_profiles/job_profile_detail.html', context)


@login_required
def job_profile_admin_list(request):
    """직무기술서 관리 목록 (관리자용 - 개선)"""
    # 권한 체크
    if not request.user.is_staff:
        messages.error(request, "관리자만 접근할 수 있습니다.")
        return redirect('job_profiles:list')
    
    # 검색 및 필터
    search_query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    is_active = request.GET.get('is_active', '')
    
    # 정렬 옵션 (신규)
    sort_by = request.GET.get('sort', 'updated_at')
    order = request.GET.get('order', 'desc')
    
    # 기본 쿼리셋
    profiles = JobProfile.objects.select_related(
        'job_role__job_type__category',
        'created_by',
        'updated_by'
    )
    
    # 검색어 필터
    if search_query:
        profiles = profiles.filter(
            Q(job_role__name__icontains=search_query) |
            Q(job_role__code__icontains=search_query)
        )
    
    # 직군 필터
    if category_id:
        profiles = profiles.filter(job_role__job_type__category_id=category_id)
    
    # 활성화 필터
    if is_active:
        profiles = profiles.filter(is_active=(is_active == '1'))
    
    # 정렬 적용
    order_prefix = '-' if order == 'desc' else ''
    if sort_by == 'name':
        profiles = profiles.order_by(f'{order_prefix}job_role__name')
    elif sort_by == 'created_at':
        profiles = profiles.order_by(f'{order_prefix}created_at')
    else:  # updated_at
        profiles = profiles.order_by(f'{order_prefix}updated_at')
    
    # 통계
    total_count = profiles.count()
    active_count = profiles.filter(is_active=True).count()
    
    # 일괄 작업용 선택된 항목 (신규)
    selected_ids = request.GET.getlist('selected')
    
    # 페이지네이션
    paginator = Paginator(profiles, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 필터용 데이터
    categories = JobCategory.objects.all()
    
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_is_active': is_active,
        'total_count': total_count,
        'active_count': active_count,
        'sort_by': sort_by,  # 신규
        'order': order,  # 신규
        'selected_ids': selected_ids,  # 신규
    }
    
    return render(request, 'job_profiles/job_profile_admin_list.html', context)


@login_required
@csrf_exempt
def bulk_update_status(request):
    """일괄 상태 변경 (신규)"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'message': '권한이 없습니다.'})
    
    if request.method == 'POST':
        data = json.loads(request.body)
        profile_ids = data.get('profile_ids', [])
        action = data.get('action', '')  # activate, deactivate
        
        if not profile_ids:
            return JsonResponse({'success': False, 'message': '선택된 항목이 없습니다.'})
        
        try:
            profiles = JobProfile.objects.filter(id__in=profile_ids)
            
            if action == 'activate':
                updated = profiles.update(is_active=True, updated_at=timezone.now())
                message = f'{updated}개 항목이 활성화되었습니다.'
            elif action == 'deactivate':
                updated = profiles.update(is_active=False, updated_at=timezone.now())
                message = f'{updated}개 항목이 비활성화되었습니다.'
            else:
                return JsonResponse({'success': False, 'message': '잘못된 작업입니다.'})
            
            return JsonResponse({'success': True, 'message': message, 'updated': updated})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': '잘못된 요청입니다.'})
