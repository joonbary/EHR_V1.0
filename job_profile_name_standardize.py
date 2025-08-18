#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
직무명 자동 표준화 및 유사명/오타 자동 교정 시스템
- Levenshtein 거리, Jaro-Winkler 등 ML 기반 유사도 매칭
- 표준 직무명과 자동 매핑 및 교정
- 불일치/유사어 로깅 및 관리자 검증
"""

import os
import sys
import io
import json
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Set
import logging
from dataclasses import dataclass, asdict
import Levenshtein
from difflib import SequenceMatcher
import re
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 한글 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


@dataclass
class MappingResult:
    """직무명 매핑 결과"""
    original: str
    standardized: str
    confidence: float
    method: str
    similarity_scores: Dict[str, float]
    is_exact_match: bool
    requires_review: bool
    suggestions: List[str]


class JobNameStandardizer:
    """직무명 표준화 엔진"""
    
    def __init__(self, standard_job_map: Dict[str, List[str]]):
        """
        Args:
            standard_job_map: 표준 직무명 매핑 (직종: [직무명 리스트])
        """
        self.standard_job_map = standard_job_map
        self.all_standard_jobs = self._flatten_job_list()
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # 유사어/동의어 사전
        self.synonym_map = {
            '모기지사업': '모기지기획',
            '모기지영업': '모기지기획',
            'NPL영업기획': 'NPL사업기획',
            'PL기획': '개인신용대출기획',
            '플랫폼/핀테크': '디지털플랫폼',
            '데이터/통계': '데이터분석',
            'HRM': '인사관리',
            'HRD': '인재개발',
            'PR': '홍보',
            'IB금융': '투자금융',
            '여신영업': '기업여신영업',
            '수신영업': '예금영업',
            '수신기획': '예금기획',
            '수신고객지원': '예금고객지원',
            '여신고객지원': '대출고객지원',
            '채권관리지원': '채권관리',
            '사무지원': '업무지원'
        }
        
        # 직무명 정규화 패턴
        self.normalization_patterns = [
            (r'\s+', ''),  # 공백 제거
            (r'[/·]', ''),  # 특수문자 통일
            (r'팀$', ''),  # '팀' 접미사 제거
            (r'부$', ''),  # '부' 접미사 제거
            (r'관리자$', '관리'),  # 관리자 -> 관리
            (r'담당자?$', ''),  # 담당/담당자 제거
        ]
    
    def _flatten_job_list(self) -> List[str]:
        """모든 표준 직무명을 단일 리스트로 변환"""
        all_jobs = []
        for job_type, jobs in self.standard_job_map.items():
            all_jobs.extend(jobs)
        return list(set(all_jobs))
    
    def normalize_job_name(self, job_name: str) -> str:
        """직무명 정규화"""
        normalized = job_name.strip()
        
        # 정규화 패턴 적용
        for pattern, replacement in self.normalization_patterns:
            normalized = re.sub(pattern, replacement, normalized)
        
        return normalized
    
    def calculate_similarity(self, name1: str, name2: str) -> Dict[str, float]:
        """다양한 유사도 메트릭 계산"""
        # 정규화
        norm1 = self.normalize_job_name(name1)
        norm2 = self.normalize_job_name(name2)
        
        # Levenshtein 거리 기반 유사도
        lev_distance = Levenshtein.distance(norm1, norm2)
        max_len = max(len(norm1), len(norm2))
        lev_similarity = 1 - (lev_distance / max_len) if max_len > 0 else 0
        
        # Jaro-Winkler 유사도
        jaro_winkler = Levenshtein.jaro_winkler(norm1, norm2)
        
        # SequenceMatcher 유사도
        seq_similarity = SequenceMatcher(None, norm1, norm2).ratio()
        
        # 한글 자모 분해 후 유사도 (더 정밀한 한글 비교)
        jamo_similarity = self._calculate_jamo_similarity(norm1, norm2)
        
        # 토큰 기반 유사도 (단어 단위 비교)
        token_similarity = self._calculate_token_similarity(name1, name2)
        
        return {
            'levenshtein': lev_similarity,
            'jaro_winkler': jaro_winkler,
            'sequence': seq_similarity,
            'jamo': jamo_similarity,
            'token': token_similarity,
            'weighted_average': (
                lev_similarity * 0.2 +
                jaro_winkler * 0.3 +
                seq_similarity * 0.2 +
                jamo_similarity * 0.15 +
                token_similarity * 0.15
            )
        }
    
    def _calculate_jamo_similarity(self, str1: str, str2: str) -> float:
        """한글 자모 분해 후 유사도 계산"""
        def decompose_korean(text):
            result = []
            for char in text:
                if '가' <= char <= '힣':
                    code = ord(char) - 0xAC00
                    jong = code % 28
                    jung = ((code - jong) // 28) % 21
                    cho = ((code - jong) // 28 - jung) // 21
                    result.extend([cho, jung, jong])
                else:
                    result.append(ord(char))
            return result
        
        jamo1 = decompose_korean(str1)
        jamo2 = decompose_korean(str2)
        
        if not jamo1 or not jamo2:
            return 0.0
        
        matches = sum(1 for a, b in zip(jamo1, jamo2) if a == b)
        return matches / max(len(jamo1), len(jamo2))
    
    def _calculate_token_similarity(self, str1: str, str2: str) -> float:
        """토큰 기반 유사도 계산"""
        # 단어 분리 (한글, 영문, 숫자 기준)
        pattern = r'[가-힣]+|[A-Za-z]+|[0-9]+'
        tokens1 = set(re.findall(pattern, str1.lower()))
        tokens2 = set(re.findall(pattern, str2.lower()))
        
        if not tokens1 or not tokens2:
            return 0.0
        
        intersection = tokens1 & tokens2
        union = tokens1 | tokens2
        
        return len(intersection) / len(union)
    
    def find_best_match(self, job_name: str) -> MappingResult:
        """최적의 표준 직무명 매칭"""
        # 1. 동의어 사전 확인
        if job_name in self.synonym_map:
            standardized = self.synonym_map[job_name]
            return MappingResult(
                original=job_name,
                standardized=standardized,
                confidence=1.0,
                method='synonym_dictionary',
                similarity_scores={'exact_synonym': 1.0},
                is_exact_match=True,
                requires_review=False,
                suggestions=[]
            )
        
        # 2. 정확히 일치하는 경우
        if job_name in self.all_standard_jobs:
            return MappingResult(
                original=job_name,
                standardized=job_name,
                confidence=1.0,
                method='exact_match',
                similarity_scores={'exact': 1.0},
                is_exact_match=True,
                requires_review=False,
                suggestions=[]
            )
        
        # 3. 유사도 기반 매칭
        best_matches = []
        for standard_job in self.all_standard_jobs:
            scores = self.calculate_similarity(job_name, standard_job)
            best_matches.append({
                'job': standard_job,
                'score': scores['weighted_average'],
                'scores': scores
            })
        
        # 점수 순으로 정렬
        best_matches.sort(key=lambda x: x['score'], reverse=True)
        top_matches = best_matches[:5]
        
        # 최고 매치 선택
        best = top_matches[0]
        confidence = best['score']
        
        # 신뢰도 임계값에 따른 처리
        if confidence >= 0.9:
            method = 'high_confidence_match'
            requires_review = False
        elif confidence >= 0.7:
            method = 'medium_confidence_match'
            requires_review = True
        else:
            method = 'low_confidence_match'
            requires_review = True
        
        return MappingResult(
            original=job_name,
            standardized=best['job'] if confidence >= 0.7 else job_name,
            confidence=confidence,
            method=method,
            similarity_scores=best['scores'],
            is_exact_match=False,
            requires_review=requires_review,
            suggestions=[m['job'] for m in top_matches]
        )


class JobProfileStandardizer:
    """직무 프로필 표준화 처리기"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # 표준 직무명 정의 (OK금융그룹 기준)
        self.standard_jobs = {
            'IT기획': ['시스템기획', '정보전략기획', 'IT기획'],
            'IT개발': ['시스템개발', '애플리케이션개발', '솔루션개발'],
            'IT운영': ['시스템관리', '서비스운영', 'IT운영관리', '인프라관리'],
            '경영관리': [
                '감사', '인사관리', '인재개발', '경영지원', '비서', '홍보',
                '경영기획', '디자인', '리스크관리', '마케팅', '스포츠사무관리',
                '자금', '재무회계', '정보보안', '준법지원', '총무'
            ],
            '투자금융': ['투자금융', 'IB업무', '기업금융자문'],
            '기업금융': ['기업영업기획', '기업여신심사', '기업여신관리'],
            '기업영업': ['기업여신영업', '기업고객관리'],
            '리테일금융': [
                '데이터분석', '디지털플랫폼', 'NPL사업기획', '리테일심사기획',
                '개인신용대출기획', '모기지기획', '예금기획', '예금영업'
            ],
            '고객지원': [
                '대출고객지원', '업무지원', '예금고객지원', '채권관리'
            ]
        }
        
        self.standardizer = JobNameStandardizer(self.standard_jobs)
        self.mapping_results = []
    
    def load_job_profiles(self, excel_path: str = None, json_path: str = None) -> pd.DataFrame:
        """직무 프로필 데이터 로드"""
        if excel_path and os.path.exists(excel_path):
            self.logger.info(f"Excel 파일 로드: {excel_path}")
            df = pd.read_excel(excel_path, sheet_name='전체직무(37개)')
            return df
        elif json_path and os.path.exists(json_path):
            self.logger.info(f"JSON 파일 로드: {json_path}")
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return pd.DataFrame(data.get('jobs', []))
        else:
            raise FileNotFoundError("Excel 또는 JSON 파일을 찾을 수 없습니다.")
    
    def standardize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """데이터프레임 직무명 표준화"""
        self.logger.info("직무명 표준화 시작...")
        
        # 표준화 결과 저장
        standardized_df = df.copy()
        self.mapping_results = []
        
        # 직무명 컬럼 확인
        job_column = None
        for col in ['직무명', '직무', 'job_title']:
            if col in df.columns:
                job_column = col
                break
        
        if not job_column:
            raise ValueError("직무명 컬럼을 찾을 수 없습니다.")
        
        # 각 직무명 표준화
        for idx, row in tqdm(df.iterrows(), total=len(df), desc="직무명 표준화"):
            original_name = row[job_column]
            if pd.isna(original_name):
                continue
            
            # 표준화 수행
            result = self.standardizer.find_best_match(original_name)
            self.mapping_results.append(result)
            
            # 결과 적용
            standardized_df.at[idx, f'{job_column}_원본'] = original_name
            standardized_df.at[idx, job_column] = result.standardized
            standardized_df.at[idx, '표준화_신뢰도'] = result.confidence
            standardized_df.at[idx, '표준화_방법'] = result.method
            standardized_df.at[idx, '검토필요'] = 'Y' if result.requires_review else 'N'
            
            # 로깅
            if result.requires_review:
                self.logger.warning(
                    f"검토 필요: '{original_name}' -> '{result.standardized}' "
                    f"(신뢰도: {result.confidence:.2f})"
                )
        
        return standardized_df
    
    def generate_mapping_report(self) -> pd.DataFrame:
        """매핑 결과 상세 리포트 생성"""
        report_data = []
        
        for result in self.mapping_results:
            report_data.append({
                '원본 직무명': result.original,
                '표준 직무명': result.standardized,
                '신뢰도': f"{result.confidence:.2%}",
                '매칭 방법': result.method,
                '정확 일치': 'Y' if result.is_exact_match else 'N',
                '검토 필요': 'Y' if result.requires_review else 'N',
                'Levenshtein 유사도': f"{result.similarity_scores.get('levenshtein', 0):.2%}",
                'Jaro-Winkler 유사도': f"{result.similarity_scores.get('jaro_winkler', 0):.2%}",
                '추천 후보 1': result.suggestions[0] if len(result.suggestions) > 0 else '',
                '추천 후보 2': result.suggestions[1] if len(result.suggestions) > 1 else '',
                '추천 후보 3': result.suggestions[2] if len(result.suggestions) > 2 else ''
            })
        
        return pd.DataFrame(report_data)
    
    def export_results(self, standardized_df: pd.DataFrame, mapping_report_df: pd.DataFrame):
        """표준화 결과 내보내기"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 1. 표준화된 Excel 저장
        excel_dir = os.path.join(self.output_dir, 'standardized')
        os.makedirs(excel_dir, exist_ok=True)
        
        excel_path = os.path.join(excel_dir, f'직무프로필_표준화_{timestamp}.xlsx')
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # 표준화 데이터
            standardized_df.to_excel(writer, sheet_name='표준화_직무프로필', index=False)
            
            # 매핑 리포트
            mapping_report_df.to_excel(writer, sheet_name='매핑_상세결과', index=False)
            
            # 검토 필요 항목
            review_needed = mapping_report_df[mapping_report_df['검토 필요'] == 'Y']
            if not review_needed.empty:
                review_needed.to_excel(writer, sheet_name='검토필요항목', index=False)
            
            # 통계 요약
            stats_df = self._generate_statistics(standardized_df, mapping_report_df)
            stats_df.to_excel(writer, sheet_name='표준화_통계', index=False)
        
        # Excel 스타일 적용
        self._apply_excel_styles(excel_path)
        
        # 2. JSON 저장
        json_dir = os.path.join(self.output_dir, 'json')
        os.makedirs(json_dir, exist_ok=True)
        
        json_path = os.path.join(json_dir, f'standardized_profiles_{timestamp}.json')
        
        json_data = {
            'metadata': {
                'standardized_at': datetime.now().isoformat(),
                'total_jobs': len(standardized_df),
                'exact_matches': len([r for r in self.mapping_results if r.is_exact_match]),
                'review_needed': len([r for r in self.mapping_results if r.requires_review])
            },
            'profiles': standardized_df.to_dict('records'),
            'mapping_results': [asdict(r) for r in self.mapping_results]
        }
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        # 3. 매핑 테이블 CSV
        mapping_path = os.path.join(self.output_dir, f'job_mapping_table_{timestamp}.csv')
        mapping_report_df.to_csv(mapping_path, encoding='utf-8-sig', index=False)
        
        # 4. 로그 파일
        log_dir = os.path.join(self.output_dir, 'logs')
        os.makedirs(log_dir, exist_ok=True)
        
        log_path = os.path.join(log_dir, f'standardization_log_{timestamp}.txt')
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write(f"직무명 표준화 로그\n")
            f.write(f"처리 일시: {datetime.now()}\n")
            f.write(f"총 처리 건수: {len(self.mapping_results)}\n\n")
            
            f.write("=== 검토 필요 항목 ===\n")
            for result in self.mapping_results:
                if result.requires_review:
                    f.write(f"'{result.original}' -> '{result.standardized}' "
                           f"(신뢰도: {result.confidence:.2%})\n")
        
        return excel_path, json_path, mapping_path, log_path
    
    def _generate_statistics(self, standardized_df: pd.DataFrame, 
                           mapping_report_df: pd.DataFrame) -> pd.DataFrame:
        """표준화 통계 생성"""
        stats = []
        
        # 전체 통계
        total = len(self.mapping_results)
        exact_matches = len([r for r in self.mapping_results if r.is_exact_match])
        high_confidence = len([r for r in self.mapping_results if r.confidence >= 0.9])
        medium_confidence = len([r for r in self.mapping_results if 0.7 <= r.confidence < 0.9])
        low_confidence = len([r for r in self.mapping_results if r.confidence < 0.7])
        review_needed = len([r for r in self.mapping_results if r.requires_review])
        
        stats.append({
            '항목': '전체 직무',
            '건수': total,
            '비율': '100%'
        })
        stats.append({
            '항목': '정확 일치',
            '건수': exact_matches,
            '비율': f"{exact_matches/total*100:.1f}%"
        })
        stats.append({
            '항목': '높은 신뢰도 (90% 이상)',
            '건수': high_confidence,
            '비율': f"{high_confidence/total*100:.1f}%"
        })
        stats.append({
            '항목': '중간 신뢰도 (70-90%)',
            '건수': medium_confidence,
            '비율': f"{medium_confidence/total*100:.1f}%"
        })
        stats.append({
            '항목': '낮은 신뢰도 (70% 미만)',
            '건수': low_confidence,
            '비율': f"{low_confidence/total*100:.1f}%"
        })
        stats.append({
            '항목': '검토 필요',
            '건수': review_needed,
            '비율': f"{review_needed/total*100:.1f}%"
        })
        
        # 매칭 방법별 통계
        method_counts = {}
        for result in self.mapping_results:
            method = result.method
            method_counts[method] = method_counts.get(method, 0) + 1
        
        for method, count in method_counts.items():
            stats.append({
                '항목': f'매칭 방법: {method}',
                '건수': count,
                '비율': f"{count/total*100:.1f}%"
            })
        
        return pd.DataFrame(stats)
    
    def _apply_excel_styles(self, excel_path: str):
        """Excel 스타일 적용"""
        wb = openpyxl.load_workbook(excel_path)
        
        # 스타일 정의
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        review_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        low_conf_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 헤더 스타일
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 조건부 서식
            if sheet_name in ['표준화_직무프로필', '매핑_상세결과']:
                for row in ws.iter_rows(min_row=2):
                    # 검토 필요 항목 하이라이트
                    for cell in row:
                        if cell.column_letter in ['F', 'G']:  # 검토필요 컬럼
                            if cell.value == 'Y':
                                for c in row:
                                    c.fill = review_fill
                        
                        # 낮은 신뢰도 하이라이트
                        if '신뢰도' in str(ws.cell(1, cell.column).value):
                            try:
                                conf_value = float(str(cell.value).rstrip('%')) / 100
                                if conf_value < 0.7:
                                    cell.fill = low_conf_fill
                            except:
                                pass
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)


def generate_sample_code():
    """샘플 사용 코드 생성"""
    sample_code = """
# 직무명 표준화 사용 예제

from job_profile_name_standardize import JobProfileStandardizer, JobNameStandardizer

# 1. 기본 사용법
standardizer = JobProfileStandardizer(output_dir='./output')

# Excel 파일에서 로드 및 표준화
df = standardizer.load_job_profiles(
    excel_path='./job_profile_complete/excel/OK금융그룹_직무기술서_전체.xlsx'
)
standardized_df = standardizer.standardize_dataframe(df)

# 결과 저장
mapping_report = standardizer.generate_mapping_report()
excel_path, json_path, mapping_path, log_path = standardizer.export_results(
    standardized_df, mapping_report
)

# 2. 개별 직무명 표준화
job_standardizer = JobNameStandardizer({
    'IT기획': ['시스템기획', 'IT기획'],
    'IT개발': ['시스템개발', '애플리케이션개발']
})

# 단일 직무명 매칭
result = job_standardizer.find_best_match('시스템 기획')
print(f"원본: {result.original}")
print(f"표준화: {result.standardized}")
print(f"신뢰도: {result.confidence:.2%}")
print(f"검토필요: {result.requires_review}")

# 3. 유사도 계산
scores = job_standardizer.calculate_similarity('모기지사업', '모기지기획')
print(f"Levenshtein: {scores['levenshtein']:.2%}")
print(f"Jaro-Winkler: {scores['jaro_winkler']:.2%}")

# 4. 배치 처리
job_names = ['HRM', 'HRD', '플랫폼/핀테크', 'IB금융']
results = [job_standardizer.find_best_match(name) for name in job_names]

# 5. Django 모델 통합 예제
from job_profiles.models import JobProfile

# 표준화 후 DB 업데이트
for index, row in standardized_df.iterrows():
    try:
        job_profile = JobProfile.objects.get(id=row['id'])
        job_profile.job_title = row['직무명']  # 표준화된 직무명
        job_profile.original_title = row['직무명_원본']  # 원본 보관
        job_profile.save()
    except JobProfile.DoesNotExist:
        pass
"""
    
    return sample_code


def main():
    """메인 실행 함수"""
    # 설정
    excel_path = r"C:/Users/apro/OneDrive/Desktop/설명회자료/job_profile_complete/excel/OK금융그룹_직무기술서_전체_20250726_183527.xlsx"
    json_path = r"C:/Users/apro/OneDrive/Desktop/설명회자료/job_profile_complete/json/ok_job_profiles_complete.json"
    output_dir = r"C:/Users/apro/OneDrive/Desktop/설명회자료/job_profile_standardized"
    
    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    # 로깅 설정
    log_file = os.path.join(output_dir, 'standardization.log')
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    logger = logging.getLogger(__name__)
    logger.info("="*60)
    logger.info("직무명 자동 표준화 시스템 시작")
    logger.info("="*60)
    
    try:
        # 표준화 처리기 생성
        standardizer = JobProfileStandardizer(output_dir)
        
        # 데이터 로드
        df = standardizer.load_job_profiles(excel_path=excel_path)
        logger.info(f"총 {len(df)}개 직무 프로필 로드 완료")
        
        # 표준화 수행
        standardized_df = standardizer.standardize_dataframe(df)
        
        # 매핑 리포트 생성
        mapping_report = standardizer.generate_mapping_report()
        
        # 결과 내보내기
        excel_out, json_out, mapping_out, log_out = standardizer.export_results(
            standardized_df, mapping_report
        )
        
        # 통계 출력
        total = len(standardizer.mapping_results)
        exact = len([r for r in standardizer.mapping_results if r.is_exact_match])
        review = len([r for r in standardizer.mapping_results if r.requires_review])
        
        logger.info("="*60)
        logger.info("표준화 완료!")
        logger.info(f"총 처리: {total}개")
        logger.info(f"정확 일치: {exact}개 ({exact/total*100:.1f}%)")
        logger.info(f"검토 필요: {review}개 ({review/total*100:.1f}%)")
        logger.info("="*60)
        
        print(f"\n✅ 직무명 표준화 완료!")
        print(f"📊 처리 결과: {total}개 직무명 표준화")
        print(f"🎯 정확 일치: {exact}개")
        print(f"⚠️  검토 필요: {review}개")
        print(f"\n📁 출력 파일:")
        print(f"  - Excel: {os.path.basename(excel_out)}")
        print(f"  - JSON: {os.path.basename(json_out)}")
        print(f"  - 매핑표: {os.path.basename(mapping_out)}")
        print(f"  - 로그: {os.path.basename(log_out)}")
        
        # 샘플 코드 생성
        sample_code_path = os.path.join(output_dir, 'sample_usage.py')
        with open(sample_code_path, 'w', encoding='utf-8') as f:
            f.write(generate_sample_code())
        print(f"  - 샘플 코드: {os.path.basename(sample_code_path)}")
        
    except Exception as e:
        logger.error(f"오류 발생: {str(e)}", exc_info=True)
        print(f"\n❌ 오류 발생: {str(e)}")


if __name__ == '__main__':
    main()