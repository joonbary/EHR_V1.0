import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
import pandas as pd
from io import BytesIO

class ExcelReportGenerator:
    """Excel 리포트 생성기"""
    
    def __init__(self, title="HR Report"):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = title
        self.current_row = 1
        
        # 스타일 정의
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def add_title(self, title, subtitle=None):
        """보고서 제목 추가"""
        self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.ws[f'A{self.current_row}']
        cell.value = title
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal='center')
        self.current_row += 1
        
        if subtitle:
            self.ws.merge_cells(f'A{self.current_row}:H{self.current_row}')
            cell = self.ws[f'A{self.current_row}']
            cell.value = subtitle
            cell.font = Font(size=12, italic=True)
            cell.alignment = Alignment(horizontal='center')
            self.current_row += 1
            
        self.current_row += 1  # 빈 줄
        
    def add_headers(self, headers):
        """테이블 헤더 추가"""
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        self.current_row += 1
        
    def add_data_row(self, data):
        """데이터 행 추가"""
        for col, value in enumerate(data, 1):
            cell = self.ws.cell(row=self.current_row, column=col)
            cell.value = value
            cell.border = self.border
            
            # 숫자 포맷팅
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0'
        self.current_row += 1
        
    def auto_adjust_columns(self):
        """열 너비 자동 조정"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
            
    def save_to_response(self, filename="report.xlsx"):
        """HTTP Response로 저장"""
        self.auto_adjust_columns()
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        self.wb.save(response)
        return response 