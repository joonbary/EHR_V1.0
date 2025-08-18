#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
OK금융그룹 직무기술서 PDF 파싱 및 데이터베이스 마이그레이션 자동화 스크립트 v2

PDF 구조에 맞춘 개선된 파싱 로직
"""

import os
import sys
import io
import json
import re
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd
from tqdm import tqdm
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 한글 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 로깅 설정
def setup_logging(output_dir: str):
    """로깅 설정"""
    log_dir = os.path.join(output_dir, 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    log_file = os.path.join(log_dir, f'migration_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    return logging.getLogger(__name__)


class JobProfilePDFParser:
    """OK금융그룹 직무기술서 PDF 파서"""
    
    def __init__(self, pdf_path: str, output_dir: str):
        self.pdf_path = pdf_path
        self.output_dir = output_dir
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # OK금융그룹 직무분류체계
        self.job_categories = {
            'IT기획': {
                'id': 'it_planning',
                'category': 'IT/디지털',
                'jobs': ['시스템기획']
            },
            'IT개발': {
                'id': 'it_development',
                'category': 'IT/디지털',
                'jobs': ['시스템개발']
            },
            'IT운영': {
                'id': 'it_operation',
                'category': 'IT/디지털',
                'jobs': ['시스템관리', '서비스운영']
            },
            '경영관리': {
                'id': 'management',
                'category': '경영지원',
                'jobs': ['감사', 'HRM', 'HRD', '경영지원', '비서', 'PR', '경영기획', 
                         '디자인', '리스크관리', '마케팅', '스포츠사무관리', '자금', 
                         '재무회계', '정보보안', '준법지원', '총무']
            },
            '투자금융': {
                'id': 'investment',
                'category': '영업/마케팅',
                'jobs': ['IB금융']
            },
            '기업금융': {
                'id': 'corporate_finance',
                'category': '영업/마케팅',
                'jobs': ['기업영업기획', '기업여신심사', '기업여신관리']
            },
            '기업영업': {
                'id': 'corporate_sales',
                'category': '영업/마케팅',
                'jobs': ['여신영업']
            },
            '리테일금융': {
                'id': 'retail_finance',
                'category': '영업/마케팅',
                'jobs': ['데이터/통계', '플랫폼/핀테크', 'NPL영업기획', '리테일심사기획', 
                         'PL기획', '모기지기획', '수신기획', '수신영업']
            },
            '고객지원': {
                'id': 'customer_support',
                'category': '영업/마케팅',
                'jobs': ['여신고객지원', '사무지원', '수신고객지원', '채권관리지원']
            }
        }
        
        self.parsed_jobs = []
        self.validation_errors = []
    
    def extract_text_from_pdf(self) -> List[Tuple[int, str]]:
        """PDF에서 텍스트 추출"""
        self.logger.info(f"PDF 파일 읽기 시작: {self.pdf_path}")
        
        try:
            doc = fitz.open(self.pdf_path)
            pages_text = []
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text()
                pages_text.append((page_num + 1, text))
                
            doc.close()
            self.logger.info(f"총 {len(pages_text)}페이지 추출 완료")
            return pages_text
            
        except Exception as e:
            self.logger.error(f"PDF 읽기 오류: {str(e)}")
            raise
    
    def parse_job_from_pages(self, pages_text: List[Tuple[int, str]]) -> List[Dict]:
        """페이지에서 직무 정보 파싱"""
        jobs = []
        current_job = None
        
        for page_num, text in pages_text:
            lines = text.strip().split('\n')
            
            # 직무 시작 페이지 감지 (예: "직무: 시스템기획")
            for i, line in enumerate(lines):
                if line.startswith('직무:'):
                    # 이전 직무 저장
                    if current_job:
                        jobs.append(current_job)
                    
                    # 새 직무 시작
                    job_name = line.replace('직무:', '').strip()
                    self.logger.info(f"페이지 {page_num}: 직무 발견 - {job_name}")
                    
                    current_job = {
                        'job_title': job_name,
                        'page_number': page_num,
                        'raw_text': '',
                        'responsibilities': [],
                        'qualifications': {
                            'basic': [],
                            'advanced': []
                        }
                    }
                    
                    # 이전 페이지에서 직종 정보 찾기
                    if page_num > 1:
                        prev_text = pages_text[page_num - 2][1] if page_num >= 2 else ''
                        job_type_match = re.search(r'직종([^\n]+)', prev_text)
                        if job_type_match:
                            job_type = job_type_match.group(1).strip()
                            current_job['job_type'] = job_type
                            
                            # 직군 매핑
                            for type_key, type_info in self.job_categories.items():
                                if type_key in job_type:
                                    current_job['job_category'] = type_info['category']
                                    current_job['category_id'] = type_info['id']
                                    current_job['mapped_type'] = type_key
                                    break
            
            # 현재 직무가 있으면 내용 파싱
            if current_job:
                current_job['raw_text'] += text + '\n'
                
                # 핵심역할&책임 파싱
                if '핵심역할&책임' in text:
                    # 테이블 형식으로 되어 있는 역할과 책임 파싱
                    in_responsibility = False
                    for line in lines:
                        if '핵심역할&책임' in line and '정의' in line:
                            in_responsibility = True
                            continue
                        
                        if in_responsibility and line.strip():
                            # 빈 줄이나 다음 섹션 시작까지
                            if '기본기술&지식' in line or '응용기술&지식' in line:
                                in_responsibility = False
                                break
                            
                            # 역할명과 설명 분리
                            parts = line.split(maxsplit=1)
                            if len(parts) >= 2 and not any(keyword in parts[0] for keyword in ['정의', '핵심역할']):
                                role_desc = parts[1] if len(parts) > 1 else line
                                if len(role_desc) > 20:  # 의미있는 설명만
                                    current_job['responsibilities'].append(role_desc)
                
                # 기술&지식 파싱
                if '기본기술&지식' in text:
                    section_text = text[text.find('기본기술&지식'):]
                    if '응용기술&지식' in section_text:
                        basic_section = section_text[:section_text.find('응용기술&지식')]
                    else:
                        basic_section = section_text
                    
                    # 기본 기술 추출
                    skill_lines = basic_section.split('\n')
                    for line in skill_lines[1:]:  # 헤더 제외
                        if line.strip() and len(line) > 10 and '정의' not in line:
                            if not any(keyword in line for keyword in ['기본기술&지식', '응용기술&지식', '직무:']):
                                current_job['qualifications']['basic'].append(line.strip())
                
                if '응용기술&지식' in text:
                    section_text = text[text.find('응용기술&지식'):]
                    skill_lines = section_text.split('\n')
                    for line in skill_lines[1:]:  # 헤더 제외
                        if line.strip() and len(line) > 10 and '정의' not in line:
                            if not any(keyword in line for keyword in ['기본기술&지식', '응용기술&지식', '직무:', 'Job Profile']):
                                current_job['qualifications']['advanced'].append(line.strip())
        
        # 마지막 직무 저장
        if current_job:
            jobs.append(current_job)
        
        return jobs
    
    def validate_job_data(self, job_data: Dict) -> Tuple[bool, List[str]]:
        """직무 데이터 검증"""
        errors = []
        
        # 필수 필드 검증
        if not job_data.get('job_title'):
            errors.append("직무명 누락")
        
        if not job_data.get('responsibilities'):
            errors.append("역할 및 책임 누락")
        elif len(job_data['responsibilities']) < 2:
            errors.append("역할 및 책임이 2개 미만")
        
        if not job_data.get('qualifications'):
            errors.append("자격요건 누락")
        else:
            basic_skills = job_data['qualifications'].get('basic', [])
            advanced_skills = job_data['qualifications'].get('advanced', [])
            if len(basic_skills) + len(advanced_skills) < 2:
                errors.append("기술/지식 요건이 2개 미만")
        
        # 직무분류 매핑 검증
        if not job_data.get('job_category'):
            errors.append("직군 분류 누락")
        
        is_valid = len(errors) == 0
        return is_valid, errors
    
    def parse_all_jobs(self) -> List[Dict]:
        """전체 PDF 파싱 실행"""
        self.logger.info("직무기술서 파싱 시작")
        
        # PDF 텍스트 추출
        pages_text = self.extract_text_from_pdf()
        
        # 직무 정보 파싱
        jobs = self.parse_job_from_pages(pages_text)
        
        # 검증 및 정리
        with tqdm(total=len(jobs), desc="직무 검증") as pbar:
            for job in jobs:
                # 검증
                is_valid, errors = self.validate_job_data(job)
                job['is_valid'] = is_valid
                job['validation_errors'] = errors
                
                if not is_valid:
                    self.logger.warning(f"검증 실패 - {job.get('job_title', 'Unknown')}: {errors}")
                
                self.parsed_jobs.append(job)
                pbar.update(1)
        
        self.logger.info(f"총 {len(self.parsed_jobs)}개 직무 파싱 완료")
        valid_count = sum(1 for job in self.parsed_jobs if job['is_valid'])
        self.logger.info(f"유효한 직무: {valid_count}개")
        
        return self.parsed_jobs


class JobProfileMigrator:
    """직무기술서 데이터베이스 마이그레이션"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # Django 설정
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ehr_system.settings')
        import django
        django.setup()
    
    def migrate_to_database(self, jobs_data: List[Dict]) -> Dict[str, int]:
        """Django 데이터베이스로 마이그레이션"""
        from job_profiles.models import JobCategory, JobType, JobRole, JobProfile
        from django.db import transaction
        
        stats = {
            'categories_created': 0,
            'types_created': 0,
            'roles_created': 0,
            'profiles_created': 0,
            'errors': 0
        }
        
        self.logger.info("데이터베이스 마이그레이션 시작")
        
        # 직군/직종 생성
        categories_map = {}
        types_map = {}
        
        with transaction.atomic():
            # 직군 생성 (IT/디지털, 경영지원, 영업/마케팅)
            main_categories = {
                'IT/디지털': 'it_digital',
                '경영지원': 'management_support',
                '영업/마케팅': 'sales_marketing'
            }
            
            for cat_name, cat_id in main_categories.items():
                category, created = JobCategory.objects.get_or_create(
                    name=cat_name,
                    defaults={'is_active': True}
                )
                categories_map[cat_name] = category
                if created:
                    stats['categories_created'] += 1
                    self.logger.info(f"직군 생성: {cat_name}")
        
        # 직무별 처리
        with tqdm(total=len(jobs_data), desc="DB 마이그레이션") as pbar:
            for job_data in jobs_data:
                try:
                    if not job_data['is_valid']:
                        self.logger.warning(f"검증 실패로 건너뜀: {job_data.get('job_title')}")
                        stats['errors'] += 1
                        pbar.update(1)
                        continue
                    
                    with transaction.atomic():
                        # 직군 가져오기
                        category_name = job_data.get('job_category', '경영지원')
                        category = categories_map.get(category_name)
                        if not category:
                            self.logger.error(f"직군을 찾을 수 없음: {category_name}")
                            stats['errors'] += 1
                            pbar.update(1)
                            continue
                        
                        # 직종 생성
                        job_type_name = job_data.get('mapped_type', job_data.get('job_type', '기타'))
                        job_type, created = JobType.objects.get_or_create(
                            category=category,
                            name=job_type_name,
                            defaults={'is_active': True}
                        )
                        if created:
                            stats['types_created'] += 1
                            self.logger.info(f"직종 생성: {job_type_name}")
                        
                        # 직무 역할 생성
                        job_role, created = JobRole.objects.get_or_create(
                            job_type=job_type,
                            name=job_data['job_title'],
                            defaults={'is_active': True}
                        )
                        if created:
                            stats['roles_created'] += 1
                        
                        # 직무 프로필 생성
                        responsibilities_text = '\n'.join(job_data.get('responsibilities', []))
                        
                        # 자격요건 통합
                        qualifications = []
                        basic_skills = job_data.get('qualifications', {}).get('basic', [])
                        advanced_skills = job_data.get('qualifications', {}).get('advanced', [])
                        
                        if basic_skills:
                            qualifications.append('[기본 기술/지식]')
                            qualifications.extend(basic_skills)
                        
                        if advanced_skills:
                            qualifications.append('\n[응용 기술/지식]')
                            qualifications.extend(advanced_skills)
                        
                        qualifications_text = '\n'.join(qualifications)
                        
                        job_profile, created = JobProfile.objects.update_or_create(
                            job_role=job_role,
                            defaults={
                                'role_responsibility': responsibilities_text,
                                'qualification': qualifications_text,
                                'is_active': True
                            }
                        )
                        
                        if created:
                            stats['profiles_created'] += 1
                            self.logger.info(f"생성됨: {job_data['job_title']}")
                        else:
                            self.logger.info(f"업데이트됨: {job_data['job_title']}")
                
                except Exception as e:
                    self.logger.error(f"마이그레이션 오류 - {job_data.get('job_title')}: {str(e)}")
                    stats['errors'] += 1
                
                pbar.update(1)
        
        self.logger.info(f"마이그레이션 완료: {stats}")
        return stats
    
    def export_to_json(self, jobs_data: List[Dict]) -> str:
        """JSON 파일로 내보내기"""
        json_dir = os.path.join(self.output_dir, 'json')
        os.makedirs(json_dir, exist_ok=True)
        
        # raw_text 제거 (파일 크기 줄이기)
        clean_jobs = []
        for job in jobs_data:
            clean_job = {k: v for k, v in job.items() if k != 'raw_text'}
            clean_jobs.append(clean_job)
        
        # 전체 데이터
        all_jobs_file = os.path.join(json_dir, 'all_job_profiles.json')
        with open(all_jobs_file, 'w', encoding='utf-8') as f:
            json.dump(clean_jobs, f, ensure_ascii=False, indent=2)
        
        # 직군별 분리
        by_category = {}
        for job in clean_jobs:
            category = job.get('job_category', '미분류')
            if category not in by_category:
                by_category[category] = []
            by_category[category].append(job)
        
        for category, jobs in by_category.items():
            category_file = os.path.join(json_dir, f'{category.replace("/", "_")}_jobs.json')
            with open(category_file, 'w', encoding='utf-8') as f:
                json.dump(jobs, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"JSON 내보내기 완료: {json_dir}")
        return all_jobs_file
    
    def export_to_excel(self, jobs_data: List[Dict]) -> str:
        """Excel 파일로 내보내기"""
        excel_dir = os.path.join(self.output_dir, 'excel')
        os.makedirs(excel_dir, exist_ok=True)
        
        excel_file = os.path.join(excel_dir, 'ok_job_profiles.xlsx')
        
        # 데이터프레임 생성
        df_data = []
        for job in jobs_data:
            # 자격요건 정리
            basic_skills = '\n'.join(job.get('qualifications', {}).get('basic', []))
            advanced_skills = '\n'.join(job.get('qualifications', {}).get('advanced', []))
            
            df_data.append({
                '페이지': job.get('page_number', ''),
                '직군': job.get('job_category', ''),
                '직종': job.get('mapped_type', job.get('job_type', '')),
                '직무명': job.get('job_title', ''),
                '역할 및 책임': '\n'.join(job.get('responsibilities', [])),
                '기본 기술/지식': basic_skills,
                '응용 기술/지식': advanced_skills,
                '검증상태': '유효' if job.get('is_valid') else '검증실패',
                '검증오류': ', '.join(job.get('validation_errors', []))
            })
        
        df = pd.DataFrame(df_data)
        
        # Excel Writer
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 전체 데이터 시트
            df.to_excel(writer, sheet_name='전체직무', index=False)
            
            # 직군별 시트
            for category in df['직군'].unique():
                if pd.notna(category):
                    category_df = df[df['직군'] == category]
                    sheet_name = category.replace('/', '_')[:31]  # Excel 시트명 제한
                    category_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 검증 실패 시트
            invalid_df = df[df['검증상태'] == '검증실패']
            if not invalid_df.empty:
                invalid_df.to_excel(writer, sheet_name='검증실패', index=False)
        
        # 스타일 적용
        wb = openpyxl.load_workbook(excel_file)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 헤더 스타일
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 행 높이 조정 (긴 텍스트용)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value and '\n' in str(cell.value):
                        ws.row_dimensions[cell.row].height = 80
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        wb.save(excel_file)
        self.logger.info(f"Excel 내보내기 완료: {excel_file}")
        return excel_file
    
    def generate_migration_report(self, jobs_data: List[Dict], stats: Dict) -> str:
        """마이그레이션 리포트 생성"""
        report_file = os.path.join(self.output_dir, 'migration_report.md')
        
        # 통계 계산
        total_jobs = len(jobs_data)
        valid_jobs = sum(1 for job in jobs_data if job['is_valid'])
        invalid_jobs = total_jobs - valid_jobs
        
        # 직군별 통계
        category_stats = {}
        for job in jobs_data:
            category = job.get('job_category', '미분류')
            if category not in category_stats:
                category_stats[category] = {'total': 0, 'valid': 0}
            category_stats[category]['total'] += 1
            if job['is_valid']:
                category_stats[category]['valid'] += 1
        
        # 리포트 생성
        report_content = f"""# OK금융그룹 직무기술서 마이그레이션 리포트

## 📊 전체 요약
- 생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- PDF 파일: OK_Job Profile.pdf
- 총 직무 수: {total_jobs}개
- 유효 직무: {valid_jobs}개 ({valid_jobs/total_jobs*100:.1f}%)
- 검증 실패: {invalid_jobs}개

## 📈 데이터베이스 마이그레이션 결과
- 직군 생성: {stats.get('categories_created', 0)}개
- 직종 생성: {stats.get('types_created', 0)}개
- 직무 생성: {stats.get('roles_created', 0)}개
- 프로필 생성: {stats.get('profiles_created', 0)}개
- 오류 발생: {stats.get('errors', 0)}건

## 📁 직군별 현황
"""
        
        for category, stat in sorted(category_stats.items()):
            report_content += f"""
### {category}
- 총 직무: {stat['total']}개
- 유효: {stat['valid']}개 ({stat['valid']/stat['total']*100:.1f}%)
"""
        
        # 검증 실패 상세
        if invalid_jobs > 0:
            report_content += "\n## ⚠️ 검증 실패 직무\n"
            for job in jobs_data:
                if not job['is_valid']:
                    errors = ', '.join(job['validation_errors'])
                    report_content += f"- {job.get('job_title', 'Unknown')}: {errors}\n"
        
        # 파싱된 직무 목록
        report_content += "\n## 📋 파싱된 직무 목록\n"
        for i, job in enumerate(jobs_data, 1):
            status = "✅" if job['is_valid'] else "❌"
            report_content += f"{i}. {status} {job.get('job_title', 'Unknown')} "
            report_content += f"({job.get('job_category', '미분류')} > {job.get('mapped_type', '미분류')})\n"
        
        # 출력 파일 정보
        report_content += f"""
## 📤 출력 파일
- JSON: {self.output_dir}/json/all_job_profiles.json
- Excel: {self.output_dir}/excel/ok_job_profiles.xlsx
- 로그: {self.output_dir}/logs/

## ✅ 다음 단계
1. Excel 파일에서 검증 실패 직무 확인 및 수동 수정
2. 데이터베이스 백업
3. 프로덕션 환경 적용
"""
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report_content)
        
        self.logger.info(f"마이그레이션 리포트 생성: {report_file}")
        return report_file


def main():
    """메인 실행 함수"""
    # 설정
    pdf_path = r"C:/Users/apro/OneDrive/Desktop/설명회자료/OK_Job Profile.pdf"
    output_dir = r"C:/Users/apro/OneDrive/Desktop/설명회자료/job_profile_migrated"
    
    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    # 로깅 설정
    logger = setup_logging(output_dir)
    logger.info("="*60)
    logger.info("OK금융그룹 직무기술서 마이그레이션 시작")
    logger.info("="*60)
    
    try:
        # 1. PDF 파싱
        parser = JobProfilePDFParser(pdf_path, output_dir)
        jobs_data = parser.parse_all_jobs()
        
        if not jobs_data:
            logger.error("파싱된 직무가 없습니다.")
            return
        
        # 2. 데이터베이스 마이그레이션
        migrator = JobProfileMigrator(output_dir)
        
        # Django 환경 확인
        try:
            stats = migrator.migrate_to_database(jobs_data)
        except Exception as e:
            logger.warning(f"데이터베이스 마이그레이션 실패: {str(e)}")
            logger.info("JSON/Excel 내보내기만 진행합니다.")
            stats = {'errors': len(jobs_data)}
        
        # 3. JSON 내보내기
        json_file = migrator.export_to_json(jobs_data)
        
        # 4. Excel 내보내기
        excel_file = migrator.export_to_excel(jobs_data)
        
        # 5. 리포트 생성
        report_file = migrator.generate_migration_report(jobs_data, stats)
        
        # 완료 메시지
        logger.info("="*60)
        logger.info("마이그레이션 완료!")
        logger.info(f"출력 디렉토리: {output_dir}")
        logger.info(f"리포트 확인: {report_file}")
        logger.info("="*60)
        
        # 결과 출력
        print("\n✅ 마이그레이션 완료!")
        print(f"📁 출력 디렉토리: {output_dir}")
        print(f"📄 JSON: {json_file}")
        print(f"📊 Excel: {excel_file}")
        print(f"📋 리포트: {report_file}")
        
    except Exception as e:
        logger.error(f"마이그레이션 중 오류 발생: {str(e)}", exc_info=True)
        raise


if __name__ == '__main__':
    main()