#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
OK금융그룹 직무기술서 PDF 파싱 및 데이터베이스 마이그레이션 - 최종 버전

PDF 구조:
- 3페이지: 직군/직종 헤더 (예: IT기획직종시스템기획)
- 4페이지: 직무명 및 핵심역할&책임
- 5페이지: 기본기술&지식, 응용기술&지식
"""

import os
import sys
import io
import json
import re
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd
from tqdm import tqdm
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 한글 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 로깅 설정
def setup_logging(output_dir: str):
    """로깅 설정"""
    log_dir = os.path.join(output_dir, 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    log_file = os.path.join(log_dir, f'migration_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    return logging.getLogger(__name__)


class JobProfilePDFParser:
    """OK금융그룹 직무기술서 PDF 파서"""
    
    def __init__(self, pdf_path: str, output_dir: str):
        self.pdf_path = pdf_path
        self.output_dir = output_dir
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # OK금융그룹 직무분류체계
        self.job_mapping = {
            'IT기획': ('IT/디지털', 'IT기획'),
            'IT개발': ('IT/디지털', 'IT개발'),
            'IT운영': ('IT/디지털', 'IT운영'),
            '경영관리': ('경영지원', '경영관리'),
            '투자금융': ('영업/마케팅', '투자금융'),
            '기업금융': ('영업/마케팅', '기업금융'),
            '기업영업': ('영업/마케팅', '기업영업'),
            '리테일금융': ('영업/마케팅', '리테일금융'),
            '고객지원': ('영업/마케팅', '고객지원')
        }
        
        self.parsed_jobs = []
    
    def extract_text_from_pdf(self) -> List[Tuple[int, str]]:
        """PDF에서 텍스트 추출"""
        self.logger.info(f"PDF 파일 읽기 시작: {self.pdf_path}")
        
        try:
            doc = fitz.open(self.pdf_path)
            pages_text = []
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text()
                pages_text.append((page_num + 1, text))
                
            doc.close()
            self.logger.info(f"총 {len(pages_text)}페이지 추출 완료")
            return pages_text
            
        except Exception as e:
            self.logger.error(f"PDF 읽기 오류: {str(e)}")
            raise
    
    def parse_job_pages(self, pages_text: List[Tuple[int, str]]) -> List[Dict]:
        """연속된 페이지에서 직무 정보 추출"""
        jobs = []
        i = 0
        
        while i < len(pages_text):
            page_num, text = pages_text[i]
            
            # 직무 시작 페이지 찾기 (Job Profile | Non-PL직군 패턴)
            if 'Job Profile' in text and '직군' in text and '직종' in text:
                job_data = {}
                
                # 1. 헤더 페이지에서 직종 정보 추출
                lines = text.strip().split('\n')
                for line in lines:
                    # IT기획직종시스템기획 같은 패턴 찾기
                    type_match = re.search(r'([가-힣]+)직종([가-힣]+)', line)
                    if type_match:
                        job_type = type_match.group(1)
                        job_name_hint = type_match.group(2)
                        
                        # 직군/직종 매핑
                        if job_type in self.job_mapping:
                            job_data['job_category'], job_data['job_type'] = self.job_mapping[job_type]
                        else:
                            job_data['job_category'] = '기타'
                            job_data['job_type'] = job_type
                        
                        self.logger.info(f"페이지 {page_num}: 직종 발견 - {job_type}")
                
                # 2. 다음 페이지에서 직무명과 역할 추출
                if i + 1 < len(pages_text):
                    next_page_num, next_text = pages_text[i + 1]
                    
                    # 직무명 추출
                    job_match = re.search(r'직무:\s*([^\n]+)', next_text)
                    if job_match:
                        job_data['job_title'] = job_match.group(1).strip()
                        job_data['page_number'] = next_page_num
                        
                        # 핵심역할&책임 추출
                        responsibilities = self.extract_responsibilities(next_text)
                        job_data['responsibilities'] = responsibilities
                
                # 3. 그 다음 페이지에서 기술/지식 추출
                if i + 2 < len(pages_text):
                    skill_page_num, skill_text = pages_text[i + 2]
                    
                    # 기본기술&지식 추출
                    basic_skills = self.extract_basic_skills(skill_text)
                    job_data['basic_skills'] = basic_skills
                    
                    # 응용기술&지식 추출
                    advanced_skills = self.extract_advanced_skills(skill_text)
                    job_data['advanced_skills'] = advanced_skills
                
                # 유효성 검증
                if self.is_valid_job(job_data):
                    jobs.append(job_data)
                    self.logger.info(f"직무 추가: {job_data.get('job_title', 'Unknown')}")
                
                # 다음 직무로 이동 (보통 3페이지씩)
                i += 3
            else:
                i += 1
        
        return jobs
    
    def extract_responsibilities(self, text: str) -> List[str]:
        """핵심역할&책임 추출"""
        responsibilities = []
        lines = text.split('\n')
        
        # 테이블 형식의 역할 추출
        current_role = None
        current_desc = []
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            if not line or '핵심역할&책임' in line or '정의' in line:
                continue
            
            # 새로운 역할 시작 (띄어쓰기가 많거나 특정 패턴)
            if len(line.split()) <= 3 and not line[0].isdigit() and i < len(lines) - 1:
                # 이전 역할 저장
                if current_role and current_desc:
                    desc_text = ' '.join(current_desc)
                    responsibilities.append(f"{current_role}: {desc_text}")
                
                current_role = line
                current_desc = []
            else:
                # 역할 설명 추가
                if current_role and len(line) > 10:
                    current_desc.append(line)
        
        # 마지막 역할 저장
        if current_role and current_desc:
            desc_text = ' '.join(current_desc)
            responsibilities.append(f"{current_role}: {desc_text}")
        
        return responsibilities
    
    def extract_basic_skills(self, text: str) -> List[str]:
        """기본기술&지식 추출"""
        skills = []
        
        if '기본기술&지식' in text:
            start_idx = text.find('기본기술&지식')
            end_idx = text.find('응용기술&지식', start_idx) if '응용기술&지식' in text else len(text)
            
            section_text = text[start_idx:end_idx]
            lines = section_text.split('\n')
            
            current_skill = None
            current_desc = []
            
            for line in lines[1:]:  # 헤더 제외
                line = line.strip()
                
                if not line or '정의' in line or '기술&지식' in line:
                    continue
                
                # 새로운 스킬 시작
                if len(line.split()) <= 4 and not line[0].isdigit():
                    # 이전 스킬 저장
                    if current_skill and current_desc:
                        desc_text = ' '.join(current_desc)
                        skills.append(f"{current_skill}: {desc_text}")
                    
                    current_skill = line
                    current_desc = []
                else:
                    # 스킬 설명 추가
                    if current_skill and len(line) > 10:
                        current_desc.append(line)
            
            # 마지막 스킬 저장
            if current_skill and current_desc:
                desc_text = ' '.join(current_desc)
                skills.append(f"{current_skill}: {desc_text}")
        
        return skills
    
    def extract_advanced_skills(self, text: str) -> List[str]:
        """응용기술&지식 추출"""
        skills = []
        
        if '응용기술&지식' in text:
            start_idx = text.find('응용기술&지식')
            section_text = text[start_idx:]
            lines = section_text.split('\n')
            
            current_skill = None
            current_desc = []
            
            for line in lines[1:]:  # 헤더 제외
                line = line.strip()
                
                if not line or '정의' in line or 'Job Profile' in line:
                    continue
                
                # 새로운 스킬 시작
                if len(line.split()) <= 4 and not line[0].isdigit():
                    # 이전 스킬 저장
                    if current_skill and current_desc:
                        desc_text = ' '.join(current_desc)
                        skills.append(f"{current_skill}: {desc_text}")
                    
                    current_skill = line
                    current_desc = []
                else:
                    # 스킬 설명 추가
                    if current_skill and len(line) > 10:
                        current_desc.append(line)
            
            # 마지막 스킬 저장
            if current_skill and current_desc:
                desc_text = ' '.join(current_desc)
                skills.append(f"{current_skill}: {desc_text}")
        
        return skills
    
    def is_valid_job(self, job_data: Dict) -> bool:
        """직무 데이터 유효성 검증"""
        required_fields = ['job_title', 'job_category', 'job_type']
        for field in required_fields:
            if not job_data.get(field):
                return False
        
        # 최소한의 정보가 있는지 확인
        has_content = (
            len(job_data.get('responsibilities', [])) > 0 or
            len(job_data.get('basic_skills', [])) > 0 or
            len(job_data.get('advanced_skills', [])) > 0
        )
        
        return has_content
    
    def parse_all_jobs(self) -> List[Dict]:
        """전체 PDF 파싱 실행"""
        self.logger.info("직무기술서 파싱 시작")
        
        # PDF 텍스트 추출
        pages_text = self.extract_text_from_pdf()
        
        # 직무 정보 파싱
        self.parsed_jobs = self.parse_job_pages(pages_text)
        
        self.logger.info(f"총 {len(self.parsed_jobs)}개 직무 파싱 완료")
        
        return self.parsed_jobs


class JobProfileMigrator:
    """직무기술서 데이터베이스 마이그레이션"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # Django 설정
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ehr_system.settings')
        try:
            import django
            django.setup()
            self.django_available = True
        except:
            self.django_available = False
            self.logger.warning("Django 환경을 사용할 수 없습니다. JSON/Excel 출력만 진행합니다.")
    
    def migrate_to_database(self, jobs_data: List[Dict]) -> Dict[str, int]:
        """Django 데이터베이스로 마이그레이션"""
        if not self.django_available:
            return {'errors': len(jobs_data)}
        
        from job_profiles.models import JobCategory, JobType, JobRole, JobProfile
        from django.db import transaction
        
        stats = {
            'categories_created': 0,
            'types_created': 0,
            'roles_created': 0,
            'profiles_created': 0,
            'errors': 0
        }
        
        self.logger.info("데이터베이스 마이그레이션 시작")
        
        # 직군/직종 매핑
        categories_map = {}
        types_map = {}
        
        with transaction.atomic():
            # 직군 생성
            for job_data in jobs_data:
                category_name = job_data.get('job_category')
                if category_name and category_name not in categories_map:
                    category, created = JobCategory.objects.get_or_create(
                        name=category_name,
                        defaults={'is_active': True}
                    )
                    categories_map[category_name] = category
                    if created:
                        stats['categories_created'] += 1
                
                # 직종 생성
                type_name = job_data.get('job_type')
                if category_name and type_name:
                    type_key = f"{category_name}_{type_name}"
                    if type_key not in types_map:
                        job_type, created = JobType.objects.get_or_create(
                            category=categories_map[category_name],
                            name=type_name,
                            defaults={'is_active': True}
                        )
                        types_map[type_key] = job_type
                        if created:
                            stats['types_created'] += 1
        
        # 직무 프로필 생성
        with tqdm(total=len(jobs_data), desc="DB 마이그레이션") as pbar:
            for job_data in jobs_data:
                try:
                    with transaction.atomic():
                        # 직종 가져오기
                        type_key = f"{job_data.get('job_category')}_{job_data.get('job_type')}"
                        job_type = types_map.get(type_key)
                        
                        if not job_type:
                            self.logger.error(f"직종을 찾을 수 없음: {type_key}")
                            stats['errors'] += 1
                            pbar.update(1)
                            continue
                        
                        # 직무 역할 생성
                        job_role, created = JobRole.objects.get_or_create(
                            job_type=job_type,
                            name=job_data['job_title'],
                            defaults={'is_active': True}
                        )
                        if created:
                            stats['roles_created'] += 1
                        
                        # 역할 및 책임
                        responsibilities = '\n'.join(job_data.get('responsibilities', []))
                        
                        # 자격요건 (기본 + 응용)
                        qualifications = []
                        if job_data.get('basic_skills'):
                            qualifications.append('[필수 역량]')
                            qualifications.extend(job_data['basic_skills'])
                        
                        if job_data.get('advanced_skills'):
                            qualifications.append('\n[우대 역량]')
                            qualifications.extend(job_data['advanced_skills'])
                        
                        qualifications_text = '\n'.join(qualifications)
                        
                        # 직무 프로필 생성/업데이트
                        job_profile, created = JobProfile.objects.update_or_create(
                            job_role=job_role,
                            defaults={
                                'role_responsibility': responsibilities,
                                'qualification': qualifications_text,
                                'is_active': True
                            }
                        )
                        
                        if created:
                            stats['profiles_created'] += 1
                            self.logger.info(f"생성: {job_data['job_title']}")
                        else:
                            self.logger.info(f"업데이트: {job_data['job_title']}")
                
                except Exception as e:
                    self.logger.error(f"오류 - {job_data.get('job_title')}: {str(e)}")
                    stats['errors'] += 1
                
                pbar.update(1)
        
        self.logger.info(f"마이그레이션 완료: {stats}")
        return stats
    
    def export_to_json(self, jobs_data: List[Dict]) -> str:
        """JSON 파일로 내보내기"""
        json_dir = os.path.join(self.output_dir, 'json')
        os.makedirs(json_dir, exist_ok=True)
        
        # 전체 데이터
        all_jobs_file = os.path.join(json_dir, 'ok_job_profiles.json')
        with open(all_jobs_file, 'w', encoding='utf-8') as f:
            json.dump(jobs_data, f, ensure_ascii=False, indent=2)
        
        # 직군별 분리
        by_category = {}
        for job in jobs_data:
            category = job.get('job_category', '미분류')
            if category not in by_category:
                by_category[category] = []
            by_category[category].append(job)
        
        for category, jobs in by_category.items():
            category_file = os.path.join(json_dir, f'{category.replace("/", "_")}_jobs.json')
            with open(category_file, 'w', encoding='utf-8') as f:
                json.dump(jobs, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"JSON 내보내기 완료: {json_dir}")
        return all_jobs_file
    
    def export_to_excel(self, jobs_data: List[Dict]) -> str:
        """Excel 파일로 내보내기"""
        excel_dir = os.path.join(self.output_dir, 'excel')
        os.makedirs(excel_dir, exist_ok=True)
        
        excel_file = os.path.join(excel_dir, 'OK금융그룹_직무기술서.xlsx')
        
        # 데이터프레임 생성
        df_data = []
        for job in jobs_data:
            df_data.append({
                '페이지': job.get('page_number', ''),
                '직군': job.get('job_category', ''),
                '직종': job.get('job_type', ''),
                '직무명': job.get('job_title', ''),
                '핵심역할 및 책임': '\n'.join(job.get('responsibilities', [])),
                '필수역량': '\n'.join(job.get('basic_skills', [])),
                '우대역량': '\n'.join(job.get('advanced_skills', []))
            })
        
        df = pd.DataFrame(df_data)
        
        # Excel 저장
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 전체 시트
            df.to_excel(writer, sheet_name='전체직무', index=False)
            
            # 직군별 시트
            for category in df['직군'].unique():
                if pd.notna(category):
                    category_df = df[df['직군'] == category]
                    sheet_name = category.replace('/', '_')[:31]
                    category_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 스타일 적용
        wb = openpyxl.load_workbook(excel_file)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 헤더 스타일
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 열 너비 조정
            column_widths = {
                'A': 10,  # 페이지
                'B': 15,  # 직군
                'C': 20,  # 직종
                'D': 25,  # 직무명
                'E': 60,  # 역할
                'F': 50,  # 필수역량
                'G': 50   # 우대역량
            }
            
            for col, width in column_widths.items():
                if col in ws.column_dimensions:
                    ws.column_dimensions[col].width = width
            
            # 텍스트 줄바꿈
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    # 긴 텍스트가 있는 행 높이 조정
                    if cell.value and isinstance(cell.value, str) and '\n' in cell.value:
                        line_count = cell.value.count('\n') + 1
                        ws.row_dimensions[cell.row].height = min(15 * line_count, 200)
        
        wb.save(excel_file)
        self.logger.info(f"Excel 내보내기 완료: {excel_file}")
        return excel_file
    
    def generate_report(self, jobs_data: List[Dict], stats: Dict) -> str:
        """마이그레이션 리포트 생성"""
        report_file = os.path.join(self.output_dir, 'MIGRATION_REPORT.md')
        
        # 통계
        total_jobs = len(jobs_data)
        
        # 직군별 통계
        category_stats = {}
        for job in jobs_data:
            category = job.get('job_category', '미분류')
            if category not in category_stats:
                category_stats[category] = []
            category_stats[category].append(job['job_title'])
        
        # 리포트 작성
        report = f"""# OK금융그룹 직무기술서 마이그레이션 리포트

## 📊 요약
- **마이그레이션 일시**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- **총 직무 수**: {total_jobs}개
- **데이터베이스 저장**: {'완료' if stats.get('profiles_created', 0) > 0 else '미실행'}

## 📈 직군별 현황

"""
        
        for category, jobs in sorted(category_stats.items()):
            report += f"### {category} ({len(jobs)}개)\n"
            for job in sorted(jobs):
                report += f"- {job}\n"
            report += "\n"
        
        if self.django_available and stats.get('profiles_created', 0) > 0:
            report += f"""## 💾 데이터베이스 마이그레이션
- **직군 생성**: {stats.get('categories_created', 0)}개
- **직종 생성**: {stats.get('types_created', 0)}개
- **직무 생성**: {stats.get('roles_created', 0)}개
- **프로필 생성**: {stats.get('profiles_created', 0)}개
- **오류**: {stats.get('errors', 0)}건

"""
        
        report += f"""## 📁 생성된 파일
1. **JSON**: `{self.output_dir}/json/ok_job_profiles.json`
2. **Excel**: `{self.output_dir}/excel/OK금융그룹_직무기술서.xlsx`
3. **로그**: `{self.output_dir}/logs/`

## ✅ 확인사항
1. Excel 파일에서 모든 직무 정보가 올바르게 추출되었는지 확인
2. 필요시 Excel에서 수정 후 재업로드
3. 데이터베이스 백업 권장

---
*Generated by OK Job Profile Migration Tool*
"""
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report)
        
        self.logger.info(f"리포트 생성: {report_file}")
        return report_file


def main():
    """메인 실행 함수"""
    # 설정
    pdf_path = r"C:/Users/apro/OneDrive/Desktop/설명회자료/OK_Job Profile.pdf"
    output_dir = r"C:/Users/apro/OneDrive/Desktop/설명회자료/job_profile_migrated"
    
    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    # 로깅 설정
    logger = setup_logging(output_dir)
    logger.info("="*60)
    logger.info("OK금융그룹 직무기술서 마이그레이션 시작")
    logger.info("="*60)
    
    try:
        # 1. PDF 파싱
        parser = JobProfilePDFParser(pdf_path, output_dir)
        jobs_data = parser.parse_all_jobs()
        
        if not jobs_data:
            logger.error("파싱된 직무가 없습니다.")
            return
        
        # 2. 마이그레이션
        migrator = JobProfileMigrator(output_dir)
        
        # 데이터베이스 마이그레이션 시도
        try:
            stats = migrator.migrate_to_database(jobs_data)
        except Exception as e:
            logger.warning(f"DB 마이그레이션 실패: {str(e)}")
            stats = {'errors': len(jobs_data)}
        
        # 3. 파일 출력
        json_file = migrator.export_to_json(jobs_data)
        excel_file = migrator.export_to_excel(jobs_data)
        report_file = migrator.generate_report(jobs_data, stats)
        
        # 완료
        logger.info("="*60)
        logger.info("마이그레이션 완료!")
        logger.info("="*60)
        
        print(f"\n✅ OK금융그룹 직무기술서 마이그레이션 완료!")
        print(f"📊 총 {len(jobs_data)}개 직무 처리")
        print(f"📁 출력 위치: {output_dir}")
        print(f"📄 Excel: {os.path.basename(excel_file)}")
        print(f"💾 JSON: {os.path.basename(json_file)}")
        print(f"📋 리포트: {os.path.basename(report_file)}")
        
    except Exception as e:
        logger.error(f"오류 발생: {str(e)}", exc_info=True)
        print(f"\n❌ 오류 발생: {str(e)}")
        print(f"상세 내용은 로그 파일을 확인하세요: {output_dir}/logs/")


if __name__ == '__main__':
    main()