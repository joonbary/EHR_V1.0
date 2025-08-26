#!/usr/bin/env python
"""
OK저축은행 조직구조 기반 1000명 직원 데이터 생성 스크립트
OK금융그룹 인사제도 반영, 엑셀 파일로 저장
"""

import os
import sys
import json
import pandas as pd
from faker import Faker
from datetime import date, datetime, timedelta
import random
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 한국어 Faker
fake = Faker('ko_KR')

class OKSavingsBankEmployeeGenerator:
    def __init__(self):
        # OK저축은행 조직구조 (실제 금융권 구조 반영)
        self.organization_structure = {
            'OK저축은행': {
                '경영지원본부': {
                    '경영기획부': ['전략기획팀', '예산관리팀', '성과관리팀'],
                    '인사총무부': ['인사팀', '총무팀', '교육연수팀'],
                    '준법감시부': ['준법감시팀', '내부감사팀'],
                    '리스크관리부': ['신용리스크팀', '시장리스크팀', '운영리스크팀']
                },
                '영업본부': {
                    '리테일영업부': ['개인영업1팀', '개인영업2팀', '개인영업3팀', 'VIP영업팀'],
                    '기업영업부': ['중소기업팀', '중견기업팀', '대기업팀'],
                    '디지털영업부': ['온라인영업팀', '모바일영업팀', '플랫폼영업팀'],
                    '영업지원부': ['영업기획팀', '상품개발팀', '마케팅팀']
                },
                '여신본부': {
                    '여신심사부': ['개인심사팀', '기업심사1팀', '기업심사2팀'],
                    '여신관리부': ['채권관리팀', '특수채권팀', 'NPL관리팀'],
                    '신용평가부': ['개인평가팀', '기업평가팀', '모형개발팀']
                },
                'IT본부': {
                    'IT기획부': ['IT전략팀', 'IT예산팀', 'IT품질팀'],
                    '시스템개발부': ['코어뱅킹팀', '채널시스템팀', '정보계팀'],
                    '인프라운영부': ['서버운영팀', '네트워크팀', 'DB관리팀'],
                    '정보보안부': ['보안관제팀', '보안기획팀', '개인정보보호팀']
                },
                '디지털혁신본부': {
                    '디지털전략부': ['디지털기획팀', 'UX/UI팀', '고객경험팀'],
                    '데이터사업부': ['빅데이터팀', 'AI/ML팀', '데이터거버넌스팀'],
                    '핀테크사업부': ['오픈뱅킹팀', '마이데이터팀', '블록체인팀']
                }
            }
        }
        
        # OK금융그룹 직급체계 (신인사제도 반영)
        self.position_hierarchy = {
            'PL직군': [
                ('수습', 1, 0),
                ('사원', 1, 1),
                ('선임', 2, 2),
                ('책임', 3, 3),
                ('수석', 4, 5),
                ('팀장', 5, 7),
                ('부장', 6, 10)
            ],
            'Non-PL직군': [
                ('주임', 1, 0),
                ('대리', 2, 2),
                ('과장', 3, 4),
                ('차장', 4, 7),
                ('부장', 5, 10),
                ('부부장', 6, 12),
                ('상무', 7, 15),
                ('전무', 8, 18),
                ('부사장', 9, 20)
            ]
        }
        
        # 성장레벨별 연봉 범위 (단위: 만원)
        self.salary_ranges = {
            1: (3500, 4500),
            2: (4500, 5500),
            3: (5500, 7000),
            4: (7000, 9000),
            5: (9000, 11000),
            6: (11000, 13000),
            7: (13000, 16000),
            8: (16000, 20000),
            9: (20000, 25000)
        }
        
        # 평가등급 분포 (정규분포)
        self.evaluation_distribution = {
            'S': 0.05,
            'A': 0.20,
            'B': 0.50,
            'C': 0.20,
            'D': 0.05
        }
        
        # 학력 수준별 분포
        self.education_distribution = {
            '고등학교': 0.05,
            '전문대학': 0.10,
            '대학교': 0.60,
            '대학원(석사)': 0.20,
            '대학원(박사)': 0.05
        }

    def generate_employees(self, count=1000):
        """OK저축은행 직원 데이터 생성"""
        employees = []
        employee_no = 20250000
        
        print(f"[INFO] OK저축은행 {count}명 직원 데이터 생성 시작...")
        
        # 조직별 인원 배분 계산
        org_distribution = self._calculate_org_distribution(count)
        
        for headquarters, headquarters_count in org_distribution.items():
            for department, dept_count in headquarters_count.items():
                for team, team_count in dept_count.items():
                    # 팀별 직원 생성
                    for i in range(team_count):
                        employee_no += 1
                        employee = self._create_employee(
                            employee_no,
                            headquarters,
                            department,
                            team
                        )
                        employees.append(employee)
                        
                        if len(employees) % 100 == 0:
                            print(f"  [진행] {len(employees)}명 생성 완료...")
        
        # 부족한 인원은 랜덤 배치
        while len(employees) < count:
            employee_no += 1
            headquarters = random.choice(list(self.organization_structure['OK저축은행'].keys()))
            departments = self.organization_structure['OK저축은행'][headquarters]
            department = random.choice(list(departments.keys()))
            teams = departments[department]
            team = random.choice(teams)
            
            employee = self._create_employee(employee_no, headquarters, department, team)
            employees.append(employee)
        
        print(f"[SUCCESS] {len(employees)}명 직원 데이터 생성 완료")
        return employees[:count]  # 정확히 count명만 반환
    
    def _calculate_org_distribution(self, total_count):
        """조직별 인원 배분 계산"""
        distribution = {}
        
        # 본부별 비중 설정
        headquarters_weights = {
            '경영지원본부': 0.15,
            '영업본부': 0.35,
            '여신본부': 0.20,
            'IT본부': 0.20,
            '디지털혁신본부': 0.10
        }
        
        for headquarters, weight in headquarters_weights.items():
            headquarters_count = int(total_count * weight)
            departments = self.organization_structure['OK저축은행'][headquarters]
            
            # 부서별 균등 배분
            dept_count_per = headquarters_count // len(departments)
            distribution[headquarters] = {}
            
            for department, teams in departments.items():
                # 팀별 균등 배분
                team_count_per = dept_count_per // len(teams)
                distribution[headquarters][department] = {}
                
                for team in teams:
                    distribution[headquarters][department][team] = team_count_per
        
        return distribution
    
    def _create_employee(self, employee_no, headquarters, department, team):
        """개별 직원 데이터 생성"""
        # 성별 결정 (남:여 = 6:4)
        gender = random.choices(['M', 'F'], weights=[0.6, 0.4])[0]
        
        # 이름 생성
        if gender == 'M':
            name = fake.name_male()
        else:
            name = fake.name_female()
        
        # 직군 결정 (PL:Non-PL = 3:7)
        job_group = random.choices(['PL직군', 'Non-PL직군'], weights=[0.3, 0.7])[0]
        
        # 직급 및 경력 결정
        position_info = random.choice(self.position_hierarchy[job_group])
        position_name = position_info[0]
        growth_level = position_info[1]
        min_years = position_info[2]
        
        # 나이 계산 (직급에 따라)
        base_age = 25 + min_years + random.randint(0, 5)
        age = min(60, max(22, base_age))
        birth_date = date.today() - timedelta(days=age * 365 + random.randint(0, 365))
        
        # 입사일 계산
        years_of_service = min_years + random.randint(0, 3)
        hire_date = date.today() - timedelta(days=years_of_service * 365 + random.randint(-180, 180))
        
        # 학력 결정
        education = random.choices(
            list(self.education_distribution.keys()),
            weights=list(self.education_distribution.values())
        )[0]
        
        # 평가등급 결정
        evaluation = random.choices(
            list(self.evaluation_distribution.keys()),
            weights=list(self.evaluation_distribution.values())
        )[0]
        
        # 연봉 계산
        salary_range = self.salary_ranges[growth_level]
        base_salary = random.randint(salary_range[0], salary_range[1])
        
        # 평가에 따른 보너스
        bonus_rates = {'S': 1.5, 'A': 1.2, 'B': 1.0, 'C': 0.8, 'D': 0.5}
        bonus = int(base_salary * bonus_rates[evaluation] * 0.3)
        
        # 이메일 생성
        email_name = name.replace(' ', '').lower()
        email = f"{email_name}{employee_no}@oksb.co.kr"
        
        # 휴대폰 번호 생성
        phone = fake.phone_number()
        
        # 주소 생성
        address = fake.address()
        
        return {
            # 기본 정보
            '사번': employee_no,
            '이름': name,
            '영문명': f"{name.upper().replace(' ', '_')}",
            '성별': '남성' if gender == 'M' else '여성',
            '생년월일': birth_date.strftime('%Y-%m-%d'),
            '나이': age,
            
            # 조직 정보
            '회사': 'OK저축은행',
            '본부': headquarters,
            '부서': department,
            '팀': team,
            '소속경로': f"OK저축은행 > {headquarters} > {department} > {team}",
            
            # 직급 정보
            '직군': job_group,
            '직급': position_name,
            '성장레벨': growth_level,
            '직책': self._get_job_title(position_name, random.random() < 0.1),
            
            # 입사 정보
            '입사일': hire_date.strftime('%Y-%m-%d'),
            '그룹입사일': hire_date.strftime('%Y-%m-%d'),
            '근속년수': years_of_service,
            
            # 평가 정보
            '최종평가등급': evaluation,
            '전년평가등급': random.choices(
                list(self.evaluation_distribution.keys()),
                weights=list(self.evaluation_distribution.values())
            )[0],
            
            # 급여 정보
            '연봉': base_salary,
            '성과급': bonus,
            '총보수': base_salary + bonus,
            
            # 개인 정보
            '학력': education,
            '전공': self._get_major(education),
            '결혼여부': random.choices(['미혼', '기혼'], weights=[0.4, 0.6])[0],
            '자녀수': random.choices([0, 1, 2, 3], weights=[0.3, 0.3, 0.3, 0.1])[0] if age > 30 else 0,
            
            # 연락처
            '이메일': email,
            '휴대폰': phone,
            '비상연락처': fake.phone_number(),
            '주소': address,
            
            # 기타 정보
            '재직상태': '재직',
            '고용형태': '정규직',
            '직무분류': self._get_job_classification(department),
            '핵심인재여부': 'Y' if evaluation == 'S' and random.random() < 0.5 else 'N',
            '승진대상여부': 'Y' if years_of_service >= min_years + 2 and evaluation in ['S', 'A'] else 'N',
            
            # 교육 이수
            '필수교육이수율': random.randint(80, 100),
            '전문교육시간': random.randint(20, 120),
            
            # 자격증
            '보유자격증': self._get_certifications(department),
            
            # 특기사항
            '특기사항': self._get_special_notes(evaluation, years_of_service)
        }
    
    def _get_job_title(self, position, is_leader):
        """직책 결정"""
        if is_leader:
            if position in ['수석', '책임', '차장', '부장']:
                return '팀장'
            elif position in ['부부장', '상무']:
                return '부서장'
            elif position in ['전무', '부사장']:
                return '본부장'
        return ''
    
    def _get_major(self, education):
        """전공 결정"""
        majors = {
            '고등학교': ['인문계', '실업계', '특성화고'],
            '전문대학': ['경영학', '회계학', '전산학', '금융학'],
            '대학교': ['경영학', '경제학', '통계학', '컴퓨터공학', '수학', '법학', '행정학'],
            '대학원(석사)': ['경영학', 'MBA', '금융공학', '컴퓨터공학', '통계학', '경제학'],
            '대학원(박사)': ['경영학', '경제학', '금융공학', '통계학']
        }
        return random.choice(majors.get(education, ['일반']))
    
    def _get_job_classification(self, department):
        """직무분류 결정"""
        job_map = {
            '경영기획부': '경영기획',
            '인사총무부': '인사/총무',
            '준법감시부': '준법/감사',
            '리스크관리부': '리스크관리',
            '리테일영업부': '리테일영업',
            '기업영업부': '기업영업',
            '디지털영업부': '디지털영업',
            '영업지원부': '영업지원',
            '여신심사부': '여신심사',
            '여신관리부': '여신관리',
            '신용평가부': '신용평가',
            'IT기획부': 'IT기획',
            '시스템개발부': 'IT개발',
            '인프라운영부': 'IT운영',
            '정보보안부': '정보보안',
            '디지털전략부': '디지털전략',
            '데이터사업부': '데이터분석',
            '핀테크사업부': '핀테크'
        }
        return job_map.get(department, '일반사무')
    
    def _get_certifications(self, department):
        """부서별 관련 자격증"""
        cert_map = {
            '여신심사부': ['여신심사역', '신용분석사', 'CRA'],
            '리스크관리부': ['FRM', '국제공인리스크관리사', '신용위험분석사'],
            'IT기획부': ['정보처리기사', 'PMP', 'ITIL'],
            '시스템개발부': ['정보처리기사', 'SQLD', 'AWS인증'],
            '정보보안부': ['정보보안기사', 'CISSP', 'CISA'],
            '데이터사업부': ['데이터분석전문가', 'SQLD', 'SAS인증'],
            '리테일영업부': ['자산관리사', '펀드투자상담사', '증권투자상담사'],
            '기업영업부': ['외환전문역', '신용분석사', '여신심사역']
        }
        
        certs = cert_map.get(department, ['은행텔러', '전산회계'])
        return ', '.join(random.sample(certs, min(2, len(certs))))
    
    def _get_special_notes(self, evaluation, years):
        """특기사항 생성"""
        notes = []
        
        if evaluation == 'S':
            notes.append('우수직원 선정')
        if years >= 10:
            notes.append('장기근속자')
        if random.random() < 0.1:
            notes.append('사내강사')
        if random.random() < 0.05:
            notes.append('혁신아이디어 수상')
        
        return ', '.join(notes) if notes else ''
    
    def save_to_excel(self, employees, filename='OK저축은행_직원명부_1000명.xlsx'):
        """엑셀 파일로 저장 (서식 포함)"""
        print(f"\n[INFO] 엑셀 파일 생성 중: {filename}")
        
        # DataFrame 생성
        df = pd.DataFrame(employees)
        
        # 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "직원명부"
        
        # 헤더 스타일 설정
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 데이터 추가
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 헤더 서식 적용
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 데이터 서식 및 열 너비 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if cell.row > 1:  # 데이터 행
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                        # 숫자 필드 우측 정렬
                        if cell.column_letter in ['E', 'F', 'M', 'Q', 'R', 'S', 'V']:  # 나이, 레벨, 년수, 급여
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            if cell.column_letter in ['Q', 'R', 'S']:  # 급여 필드
                                cell.number_format = '#,##0"만원"'
                    
                    # 열 너비 자동 조정
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 행 높이 설정
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 20
        
        # 첫 번째 행(헤더) 고정
        ws.freeze_panes = 'A2'
        
        # 자동 필터 추가
        ws.auto_filter.ref = ws.dimensions
        
        # 통계 시트 추가
        stats_ws = wb.create_sheet(title="통계")
        
        # 통계 데이터 계산
        stats_data = [
            ['구분', '내용', '값'],
            ['전체', '총 직원수', len(employees)],
            ['', '', ''],
            ['본부별 인원', '', ''],
        ]
        
        # 본부별 통계
        for headquarters in df['본부'].unique():
            count = len(df[df['본부'] == headquarters])
            stats_data.append(['', headquarters, count])
        
        stats_data.append(['', '', ''])
        stats_data.append(['직급별 인원', '', ''])
        
        # 직급별 통계
        for position in df['직급'].unique():
            count = len(df[df['직급'] == position])
            stats_data.append(['', position, count])
        
        # 통계 데이터 추가
        for row in stats_data:
            stats_ws.append(row)
        
        # 통계 시트 서식 적용
        for row in stats_ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
        
        # 파일 저장
        wb.save(filename)
        print(f"[SUCCESS] 엑셀 파일 저장 완료: {filename}")
        
        return filename
    
    def save_to_csv(self, employees, filename='OK저축은행_직원명부_1000명.csv'):
        """CSV 파일로 저장"""
        df = pd.DataFrame(employees)
        df.to_csv(filename, index=False, encoding='utf-8-sig')
        print(f"[SUCCESS] CSV 파일 저장 완료: {filename}")
        return filename

def main():
    print("=" * 70)
    print("OK저축은행 1000명 직원 데이터 생성")
    print("OK금융그룹 인사제도 기반")
    print("=" * 70)
    
    generator = OKSavingsBankEmployeeGenerator()
    
    # 1000명 직원 데이터 생성
    employees = generator.generate_employees(1000)
    
    # 엑셀 파일로 저장
    excel_file = generator.save_to_excel(employees)
    
    # CSV 파일로도 저장
    csv_file = generator.save_to_csv(employees)
    
    # 통계 출력
    print("\n" + "=" * 70)
    print("생성 완료 통계")
    print("=" * 70)
    
    df = pd.DataFrame(employees)
    
    print(f"\n[총 인원] {len(employees)}명")
    
    print("\n[본부별 인원]")
    for headquarters in df['본부'].value_counts().index:
        count = df['본부'].value_counts()[headquarters]
        print(f"  - {headquarters}: {count}명")
    
    print("\n[직급별 인원]")
    for position in df['직급'].value_counts().index[:10]:
        count = df['직급'].value_counts()[position]
        print(f"  - {position}: {count}명")
    
    print("\n[평가등급 분포]")
    for grade in ['S', 'A', 'B', 'C', 'D']:
        count = len(df[df['최종평가등급'] == grade])
        percentage = count / len(df) * 100
        print(f"  - {grade}등급: {count}명 ({percentage:.1f}%)")
    
    print("\n[파일 정보]")
    print(f"  - 엑셀 파일: {excel_file}")
    print(f"  - CSV 파일: {csv_file}")
    print("\n업로드 준비 완료!")

if __name__ == "__main__":
    main()