from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from .models import Employee
from .models_organization import OrganizationStructure, OrganizationUploadHistory, EmployeeOrganizationMapping
from .models_hr import HREmployee
from .models_workforce import WeeklyWorkforceSnapshot
from .forms import EmployeeForm
from django.views import View
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.urls import reverse
import pandas as pd
from django.core.files.storage import default_storage
from django.conf import settings
import os
from django.utils.datastructures import MultiValueDictKeyError
from django.utils.encoding import smart_str
import re
import json
from datetime import date
from utils.file_manager import ExcelFileHandler
from django.contrib.auth.decorators import login_required
from django.db.models import Q

class EmployeeListView(ListView):
    model = Employee
    template_name = 'employees/employee_list_revolutionary.html'
    context_object_name = 'employees'
    paginate_by = 20
    
    def get_queryset(self):
        # Safe queryset - only() to limit fields loaded
        queryset = Employee.objects.only('id', 'name', 'email', 'department', 'position', 'phone', 'no', 
                                         'final_department', 'current_position', 'new_position',
                                         'company', 'headquarters1', 'employment_status')
        
        # 검색어 가져오기
        search_query = self.request.GET.get('q', '')
        search_type = self.request.GET.get('search_type', 'all')
        
        # 필터링 파라미터들
        company = self.request.GET.get('company', '')
        headquarters1 = self.request.GET.get('headquarters1', '')
        position = self.request.GET.get('position', '')
        employment_status = self.request.GET.get('employment_status', '')
        
        # 검색 필터링
        if search_query:
            if search_type == 'name':
                queryset = queryset.filter(name__icontains=search_query)
            elif search_type == 'email':
                queryset = queryset.filter(email__icontains=search_query)
            elif search_type == 'department':
                queryset = queryset.filter(
                    Q(department__icontains=search_query) |
                    Q(final_department__icontains=search_query) |
                    Q(headquarters1__icontains=search_query) |
                    Q(headquarters2__icontains=search_query)
                )
            elif search_type == 'position':
                queryset = queryset.filter(
                    Q(position__icontains=search_query) |
                    Q(current_position__icontains=search_query) |
                    Q(new_position__icontains=search_query)
                )
            else:  # all
                queryset = queryset.filter(
                    Q(name__icontains=search_query) |
                    Q(email__icontains=search_query) |
                    Q(department__icontains=search_query) |
                    Q(final_department__icontains=search_query) |
                    Q(headquarters1__icontains=search_query) |
                    Q(headquarters2__icontains=search_query) |
                    Q(position__icontains=search_query) |
                    Q(current_position__icontains=search_query) |
                    Q(new_position__icontains=search_query) |
                    Q(responsibility__icontains=search_query) |
                    Q(job_type__icontains=search_query)
                )
        
        # 회사 필터링
        if company:
            queryset = queryset.filter(company=company)
        
        # 본부 필터링
        if headquarters1:
            queryset = queryset.filter(headquarters1__icontains=headquarters1)
        
        # 직급 필터링
        if position:
            queryset = queryset.filter(
                Q(current_position=position) |
                Q(new_position=position) |
                Q(position=position)
            )
        
        # 재직상태 필터링
        if employment_status:
            queryset = queryset.filter(employment_status=employment_status)
        
        # 정렬 (이름 순으로 간단히)
        return queryset.order_by('name')
    
    def get_context_data(self, **kwargs):
        # Database connection retry mechanism
        from django.db import connection, OperationalError
        import time
        
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                context = super().get_context_data(**kwargs)
                
                # 전체 직원 수
                context['total_count'] = Employee.objects.count()
                # 현재 쿼리셋의 수
                context['queryset_count'] = self.get_queryset().count()
                break  # Success, exit retry loop
                
            except OperationalError as e:
                retry_count += 1
                if retry_count >= max_retries:
                    # Final retry failed, return safe defaults
                    context = {'object_list': [], 'employees': []}
                    context['total_count'] = 0
                    context['queryset_count'] = 0
                    context['error_message'] = "데이터베이스 연결 오류가 발생했습니다. 잠시 후 다시 시도해주세요."
                    return context
                
                # Close the old connection and retry
                connection.close()
                time.sleep(0.5 * retry_count)  # Exponential backoff
                continue
        
        # 통계 데이터 추가 (상단 카드용)
        from datetime import datetime, timedelta
        
        # 부서 수 계산 (중복 제거)
        departments = set()
        try:
            for dept in Employee.objects.exclude(department__isnull=True).exclude(department='').values_list('department', flat=True):
                departments.add(dept)
            for dept in Employee.objects.exclude(final_department__isnull=True).exclude(final_department='').values_list('final_department', flat=True):
                departments.add(dept)
            context['department_count'] = len(departments)
        except Exception:
            context['department_count'] = 0
        
        # 활성 계정 수 (재직 상태)
        try:
            context['active_count'] = Employee.objects.filter(employment_status='재직').count()
        except Exception:
            context['active_count'] = Employee.objects.count()
        
        # 신규 입사 (최근 30일)
        try:
            thirty_days_ago = datetime.now().date() - timedelta(days=30)
            context['new_employee_count'] = Employee.objects.filter(hire_date__gte=thirty_days_ago).count()
        except Exception:
            context['new_employee_count'] = 0
        
        # 검색 관련 컨텍스트
        context['search_query'] = self.request.GET.get('q', '')
        context['search_type'] = self.request.GET.get('search_type', 'all')
        
        # 필터 관련 컨텍스트
        context['filter_company'] = self.request.GET.get('company', '')
        context['filter_headquarters1'] = self.request.GET.get('headquarters1', '')
        context['filter_position'] = self.request.GET.get('position', '')
        context['filter_employment_status'] = self.request.GET.get('employment_status', '')
        
        # 필터 옵션용 데이터
        try:
            context['companies'] = Employee.objects.exclude(company__isnull=True).exclude(company='').values_list('company', flat=True).distinct().order_by('company')
            context['headquarters1_list'] = Employee.objects.exclude(headquarters1__isnull=True).exclude(headquarters1='').values_list('headquarters1', flat=True).distinct().order_by('headquarters1')
            context['positions'] = Employee.objects.exclude(current_position__isnull=True).exclude(current_position='').values_list('current_position', flat=True).distinct().order_by('current_position')
            context['employment_statuses'] = Employee.objects.exclude(employment_status__isnull=True).exclude(employment_status='').values_list('employment_status', flat=True).distinct().order_by('employment_status')
        except Exception:
            context['companies'] = []
            context['headquarters1_list'] = []
            context['positions'] = []
            context['employment_statuses'] = []
        
        return context

class EmployeeDetailView(DetailView):
    model = Employee
    template_name = 'employees/employee_detail_revolutionary.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        emp = self.object
        today = date.today()
        # 근무기간 계산 (hire_date가 없을 수도 있음)
        if emp.hire_date:
            years = today.year - emp.hire_date.year
            months = today.month - emp.hire_date.month
            days = today.day - emp.hire_date.day
            if days < 0:
                months -= 1
            if months < 0:
                years -= 1
                months += 12
        else:
            years = 0
            months = 0
            days = 0
        work_period_text = f"{years}년 {months}개월"
        context['today'] = today
        context['work_period_text'] = work_period_text
        return context

class EmployeeCreateView(CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employees/employee_form_revolutionary.html'
    success_url = reverse_lazy('employees:employee_list')
    
    def form_valid(self, form):
        messages.success(self.request, '직원이 성공적으로 등록되었습니다.')
        return super().form_valid(form)

class EmployeeUpdateView(UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employees/employee_form_revolutionary.html'
    success_url = reverse_lazy('employees:employee_list')
    
    def form_valid(self, form):
        messages.success(self.request, '직원 정보가 성공적으로 수정되었습니다.')
        return super().form_valid(form)

class EmployeeDeleteView(DeleteView):
    model = Employee
    template_name = 'employees/employee_confirm_delete.html'
    success_url = reverse_lazy('employees:employee_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, '직원이 성공적으로 삭제되었습니다.')
        return super().delete(request, *args, **kwargs)

@method_decorator(csrf_exempt, name="dispatch")
class BulkUploadView(View):
    template_name = 'employees/bulk_upload.html'
    sample_headers = ['이름', '부서', '직급', '입사일', '이메일', '전화번호']

    def get(self, request):
        # 샘플 템플릿 다운로드
        if 'download' in request.GET:
            df = pd.DataFrame(columns=self.sample_headers)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=employee_bulk_upload_sample.xlsx'
            df.to_excel(response, index=False)
            return response
        return render(request, self.template_name, {'sample_headers': self.sample_headers})

    def post(self, request):
        # AJAX 요청인지 확인
        if request.headers.get('Content-Type') == 'application/json':
            return self.handle_ajax_save(request)
        
        # 일반 파일 업로드 (미리보기)
        return self.handle_preview(request)
    
    def handle_preview(self, request):
        """미리보기 처리 - DB 저장 없이 검증만"""
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'error': '파일을 선택해주세요.'}, status=400)
        
        excel_handler = ExcelFileHandler()
        
        try:
            # Excel 파일 처리 및 검증
            records, total_rows = excel_handler.process_employee_excel(file)
            df = pd.DataFrame(records)
        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'파일을 읽는 중 오류가 발생했습니다: {e}'}, status=400)

        # 미리보기(최대 10개)
        preview = df.head(10).to_dict(orient='records')
        
        # 전체 데이터
        full_data = df.to_dict(orient='records')
        
        # 검증 (전체 데이터에 대해, 미리보기는 10개까지만 오류 표시)
        errors = []
        for idx, row in df.iterrows():
            if idx >= 10:  # 10개 이후는 검증만 하고 오류는 표시 안함
                break
            row_errors = []
            name = row.get('이름', '')
            email = row.get('이메일', '')
            department = row.get('부서', '')
            position = row.get('직급', '')
            hire_date = row.get('입사일', '')
            phone = row.get('전화번호', '')
            # 필수값 체크 (NaN 포함)
            if pd.isna(name) or not str(name).strip():
                row_errors.append('이름 필수')
            if pd.isna(email) or not str(email).strip():
                row_errors.append('이메일 필수')
            if pd.isna(department) or not str(department).strip():
                row_errors.append('부서 필수')
            if pd.isna(position) or not str(position).strip():
                row_errors.append('직급 필수')
            if pd.isna(hire_date) or not str(hire_date).strip():
                row_errors.append('입사일 필수')
            # 이메일 중복 체크
            if email and not pd.isna(email) and Employee.objects.filter(email=str(email).strip()).exists():
                row_errors.append('이메일 중복')
            # 날짜 형식 체크 (YYYY-MM-DD) 및 유효성 검증
            if hire_date and not pd.isna(hire_date):
                hire_date_str = str(hire_date).strip()
                if not re.match(r'^\d{4}-\d{2}-\d{2}$', hire_date_str):
                    row_errors.append('입사일 형식 오류(YYYY-MM-DD)')
                else:
                    # 실제 유효한 날짜인지 확인
                    try:
                        from datetime import datetime
                        datetime.strptime(hire_date_str, '%Y-%m-%d')
                    except ValueError:
                        row_errors.append('입사일 형식 오류(YYYY-MM-DD)')
            if row_errors:
                errors.append({'row': idx+2, 'name': str(name), 'email': str(email), 'errors': row_errors})
        return JsonResponse({
            'columns': df.columns.tolist(),
            'preview': preview,
            'full_data': full_data,  # 전체 데이터 추가
            'errors': errors,
            'total_rows': len(df)
        })
    
    def handle_ajax_save(self, request):
        """최종 저장 처리"""
        try:
            data = json.loads(request.body)
            columns = data.get('columns', [])
            rows = data.get('rows', [])
        except json.JSONDecodeError:
            return JsonResponse({'error': '잘못된 데이터 형식입니다.'}, status=400)
        
        # 검증 및 저장
        errors = []
        success_count = 0
        fail_count = 0
        created = []
        
        for idx, row in enumerate(rows):
            row_errors = []
            
            # 기본 필수 필드
            name = str(row.get('이름', '')).strip()
            email = str(row.get('이메일', '')).strip()
            hire_date = str(row.get('입사일', '')).strip()
            
            # 선택 필드들
            phone = str(row.get('전화번호', '') or row.get('연락처', '')).strip()
            company = str(row.get('회사', '')).strip()
            headquarters1 = str(row.get('본부1', '')).strip()
            headquarters2 = str(row.get('본부2', '')).strip()
            final_department = str(row.get('최종소속', '') or row.get('부서', '')).strip()
            current_position = str(row.get('직급', '')).strip()
            responsibility = str(row.get('직책', '')).strip()
            gender = str(row.get('성별', '')).strip()
            age = row.get('나이', None)
            employment_status = str(row.get('재직상태', '재직')).strip()
            job_group = str(row.get('직군', '') or row.get('직군/계열', '')).strip()
            job_type = str(row.get('직종', '')).strip()
            
            # 필수값 체크
            if not name:
                row_errors.append('이름 필수')
            if not email:
                row_errors.append('이메일 필수')
            if not hire_date:
                row_errors.append('입사일 필수')
            
            # 이메일 중복 체크
            if email and Employee.objects.filter(email=email).exists():
                row_errors.append('이메일 중복')
            
            # 날짜 형식 체크 (YYYY-MM-DD)
            if hire_date and not re.match(r'^\d{4}-\d{2}-\d{2}$', hire_date):
                row_errors.append('입사일 형식 오류(YYYY-MM-DD)')
            
            # 성별 값 검증
            if gender and gender.upper() not in ['M', 'F', '남', '여', '남성', '여성']:
                row_errors.append('성별 형식 오류(M/F 또는 남/여)')
            
            # 나이 검증
            if age:
                try:
                    age = int(age)
                    if age < 18 or age > 70:
                        row_errors.append('나이는 18-70 사이여야 합니다')
                except (ValueError, TypeError):
                    row_errors.append('나이는 숫자여야 합니다')
                    age = None
            
            if row_errors:
                errors.append({'row': idx+2, 'name': name, 'email': email, 'errors': row_errors})
                fail_count += 1
                continue
            
            # 성별 값 변환
            if gender:
                if gender.upper() in ['M', '남', '남성']:
                    gender = 'M'
                elif gender.upper() in ['F', '여', '여성']:
                    gender = 'F'
                else:
                    gender = ''
            
            # 실제 저장
            try:
                employee_data = {
                    'name': name,
                    'email': email,
                    'hire_date': hire_date,
                    'phone': phone,
                    'company': company if company else None,
                    'headquarters1': headquarters1 if headquarters1 else None,
                    'headquarters2': headquarters2 if headquarters2 else None,
                    'final_department': final_department if final_department else None,
                    'current_position': current_position if current_position else None,
                    'responsibility': responsibility if responsibility else None,
                    'gender': gender if gender else None,
                    'age': age,
                    'employment_status': employment_status,
                    'job_group': job_group if job_group else 'Non-PL',
                    'job_type': job_type if job_type else '경영관리',
                }
                
                # 기존 필드도 호환성을 위해 설정
                employee_data['department'] = final_department if final_department else 'OPERATIONS'
                employee_data['position'] = current_position if current_position else 'STAFF'
                employee_data['new_position'] = current_position if current_position else '사원'
                
                Employee.objects.create(**employee_data)
                success_count += 1
                created.append(email)
            except Exception as e:
                errors.append({'row': idx+2, 'name': name, 'email': email, 'errors': [str(e)]})
                fail_count += 1
        
        return JsonResponse({
            'success_count': success_count,
            'fail_count': fail_count,
            'errors': errors,
            'created': created
        })

def download_template(request):
    import pandas as pd
    from io import BytesIO
    
    # 확장된 헤더 (필수 + 선택 필드)
    headers = [
        '이름', '이메일', '입사일',  # 필수 필드
        '전화번호', '회사', '본부1', '본부2', '최종소속', '직급', '직책',  # 기본 선택 필드
        '성별', '나이', '재직상태', '직군/계열', '직종'  # 추가 선택 필드
    ]
    
    # 샘플 데이터 (3명 분량)
    sample = [
        ['홍길동', 'hong@ok.co.kr', '2024-01-01', '010-1234-5678', 'OK', 'IT본부', 'AI개발본부', '개발1팀', '대리', '팀장', 'M', '35', '재직', 'Non-PL', 'IT개발'],
        ['김영희', 'kim@oci.co.kr', '2023-12-01', '010-8765-4321', 'OCI', '기획본부', '전략기획본부', '기획팀', '과장', '주임', 'F', '38', '재직', 'Non-PL', '경영관리'],
        ['박민수', 'park@okds.co.kr', '2024-03-15', '010-5555-7777', 'OKDS', '영업본부', '리테일영업본부', '고객지원팀', '사원', '', 'M', '28', '재직', 'PL', '고객지원'],
    ]
    
    df = pd.DataFrame(sample, columns=headers)
    
    # 설명용 시트 추가
    with pd.ExcelWriter(BytesIO(), engine='openpyxl') as writer:
        # 메인 데이터 시트
        df.to_excel(writer, sheet_name='직원데이터', index=False)
        
        # 설명 시트
        help_data = [
            ['컬럼명', '필수여부', '설명', '예시값'],
            ['이름', 'O', '직원 성명 (한글/영문)', '홍길동'],
            ['이메일', 'O', '회사 이메일 주소 (중복불가)', 'hong@ok.co.kr'],
            ['입사일', 'O', '입사일자 (YYYY-MM-DD 형식)', '2024-01-01'],
            ['전화번호', 'X', '휴대폰 번호', '010-1234-5678'],
            ['회사', 'X', '소속 회사', 'OK, OCI, OC, OFI, OKDS, OKH, ON, OKIP, OT, OKV, EX'],
            ['본부1', 'X', '1차 본부', 'IT본부, 기획본부, 영업본부'],
            ['본부2', 'X', '2차 본부', 'AI개발본부, 전략기획본부'],
            ['최종소속', 'X', '최종 소속 부서/팀', '개발1팀, 기획팀'],
            ['직급', 'X', '직급', '사원, 선임, 주임, 대리, 과장, 차장, 부부장, 부장, 팀장, 지점장, 본부장'],
            ['직책', 'X', '직책/역할', '팀장, 주임, 선임 등'],
            ['성별', 'X', '성별', 'M(남성), F(여성), 남, 여, 남성, 여성'],
            ['나이', 'X', '나이 (숫자)', '35'],
            ['재직상태', 'X', '현재 재직 상태', '재직, 휴직, 파견, 퇴직'],
            ['직군/계열', 'X', '직군 분류', 'PL, Non-PL'],
            ['직종', 'X', '직종 분류', 'IT개발, IT기획, IT운영, 경영관리, 기업영업, 기업금융, 리테일금융, 투자금융, 고객지원'],
        ]
        
        help_df = pd.DataFrame(help_data[1:], columns=help_data[0])
        help_df.to_excel(writer, sheet_name='입력가이드', index=False)
        
        # BytesIO 객체 생성
        output = BytesIO()
        writer.book.save(output)
        output.seek(0)
    
    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % smart_str('OK금융그룹_직원업로드_템플릿.xlsx')
    return response

def organization_chart(request):
    """Revolutionary Design 기반 조직도 페이지 v2"""
    return render(request, 'employees/organization_chart_v2.html')

def hierarchy_organization_view(request):
    """계층별 조직도 페이지 (Revolutionary Design)"""
    return render(request, 'employees/hierarchy_organization_revolutionary.html')

def hierarchy_organization_api(request):
    """계층별 조직 데이터 API"""
    level = request.GET.get('level', 'all')  # executive, department, team, member
    department = request.GET.get('department', '')
    parent_id = request.GET.get('parent_id', '')
    
    try:
        if level == 'executive':
            # 임원급 (부사장 이상)
            executives = Employee.objects.filter(
                new_position__in=['회장', '부사장', '사장', '전무', '상무'],
                employment_status='재직'
            ).select_related('manager')
            
            data = []
            for exec in executives:
                subordinates = exec.get_subordinates()
                data.append({
                    'id': exec.id,
                    'name': exec.name,
                    'position': exec.new_position,
                    'department': exec.department,
                    'job_type': exec.job_type,
                    'level': exec.growth_level,
                    'age': calculate_age(exec.hire_date) if exec.hire_date else None,
                    'email': exec.email,
                    'phone': exec.phone,
                    'profile_image': exec.profile_image.url if exec.profile_image else None,
                    'subordinate_count': subordinates.count(),
                    'subordinate_departments': list(subordinates.values_list('department', flat=True).distinct()),
                    'manager': exec.manager.name if exec.manager else None
                })
        
        elif level == 'department':
            # 부서장급 (본부장, 부장급)
            dept_heads = Employee.objects.filter(
                new_position__in=['본부장', '부장'],
                employment_status='재직'
            ).select_related('manager')
            
            data = []
            for dept in dept_heads:
                subordinates = dept.get_subordinates()
                teams = subordinates.filter(new_position__in=['팀장', '지점장']).count()
                
                data.append({
                    'id': dept.id,
                    'name': dept.name,
                    'position': dept.new_position,
                    'department': dept.department,
                    'job_type': dept.job_type,
                    'level': dept.growth_level,
                    'age': calculate_age(dept.hire_date) if dept.hire_date else None,
                    'email': dept.email,
                    'phone': dept.phone,
                    'profile_image': dept.profile_image.url if dept.profile_image else None,
                    'total_members': subordinates.count(),
                    'team_count': teams,
                    'manager': dept.manager.name if dept.manager else None,
                    'organization_name': f"{dept.department} {dept.new_position}"
                })
        
        elif level == 'team':
            # 팀장급
            team_heads = Employee.objects.filter(
                new_position__in=['팀장', '지점장'],
                employment_status='재직'
            ).select_related('manager')
            
            if department:
                team_heads = team_heads.filter(department=department)
            
            data = []
            for team in team_heads:
                subordinates = team.get_subordinates()
                
                data.append({
                    'id': team.id,
                    'name': team.name,
                    'position': team.new_position,
                    'department': team.department,
                    'job_type': team.job_type,
                    'level': team.growth_level,
                    'age': calculate_age(team.hire_date) if team.hire_date else None,
                    'email': team.email,
                    'phone': team.phone,
                    'profile_image': team.profile_image.url if team.profile_image else None,
                    'team_members': subordinates.count(),
                    'manager': team.manager.name if team.manager else None,
                    'organization_name': f"{team.department} {team.job_type}팀",
                    'members': [
                        {
                            'id': member.id,
                            'name': member.name,
                            'position': member.new_position,
                            'level': member.growth_level,
                            'job_type': member.job_type
                        } for member in subordinates
                    ]
                })
        
        elif level == 'member':
            # 특정 상사의 직속 부하들
            if parent_id:
                members = Employee.objects.filter(
                    manager_id=parent_id,
                    employment_status='재직'
                ).select_related('manager')
            else:
                # 일반 직원들 (팀장급 제외)
                members = Employee.objects.filter(
                    new_position__in=['사원', '선임', '주임', '대리', '과장', '차장'],
                    employment_status='재직'
                ).select_related('manager')
                
                if department:
                    members = members.filter(department=department)
            
            data = []
            for member in members:
                data.append({
                    'id': member.id,
                    'name': member.name,
                    'position': member.new_position,
                    'department': member.department,
                    'job_type': member.job_type,
                    'level': member.growth_level,
                    'age': calculate_age(member.hire_date) if member.hire_date else None,
                    'email': member.email,
                    'phone': member.phone,
                    'profile_image': member.profile_image.url if member.profile_image else None,
                    'manager': member.manager.name if member.manager else None,
                    'hire_date': member.hire_date.strftime('%Y-%m-%d') if member.hire_date else None
                })
        
        else:  # level == 'all'
            # 전체 조직 요약
            data = {
                'executives': Employee.objects.filter(
                    new_position__in=['회장', '부사장', '사장', '전무', '상무'],
                    employment_status='재직'
                ).count(),
                'departments': Employee.objects.filter(
                    new_position__in=['본부장', '부장'],
                    employment_status='재직'
                ).count(),
                'teams': Employee.objects.filter(
                    new_position__in=['팀장', '지점장'],
                    employment_status='재직'
                ).count(),
                'members': Employee.objects.filter(
                    new_position__in=['사원', '선임', '주임', '대리', '과장', '차장'],
                    employment_status='재직'
                ).count(),
                'total': Employee.objects.filter(employment_status='재직').count(),
                'departments_list': list(Employee.objects.filter(
                    employment_status='재직'
                ).values_list('department', flat=True).distinct())
            }
        
        return JsonResponse({
            'success': True,
            'level': level,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def calculate_age(hire_date):
    """입사일로부터 근무 연수 계산"""
    from datetime import date
    today = date.today()
    return today.year - hire_date.year

def organization_data_api(request):
    """조직도 데이터 JSON API - OrganizationStructure 테이블 사용"""
    
    # 먼저 OrganizationStructure 데이터가 있는지 확인
    try:
        from employees.models_organization import OrganizationStructure
    except ImportError:
        # OrganizationStructure 모델이 없으면 기존 Employee 테이블 사용
        return organization_data_api_legacy(request)
    
    # OrganizationStructure에 데이터가 있는지 확인
    if not OrganizationStructure.objects.filter(status='active').exists():
        # 데이터가 없으면 기존 Employee 테이블 사용
        return organization_data_api_legacy(request)
    
    def build_org_tree(parent_id=None, processed=None):
        if processed is None:
            processed = set()
        
        # 순환 참조 방지
        if parent_id and parent_id in processed:
            return []
        
        if parent_id:
            processed.add(parent_id)
        
        # 해당 부모의 하위 조직들 조회
        orgs = OrganizationStructure.objects.filter(
            parent_id=parent_id,
            status='active'
        ).order_by('sort_order', 'org_code')
        
        nodes = []
        for org in orgs:
            # 조직장 정보 가져오기
            leader_info = None
            if org.leader:
                leader_info = {
                    'name': org.leader.name,
                    'position': org.leader.new_position,
                    'email': org.leader.email,
                    'phone': org.leader.phone
                }
            
            # 해당 조직의 직원 수 계산
            employee_count = 0
            if hasattr(Employee, 'objects'):
                # 조직명으로 직원 찾기 (본부, 부서, 팀 기준)
                if org.org_level == 3:  # 본부
                    employee_count = Employee.objects.filter(
                        headquarters1=org.org_name,
                        employment_status='재직'
                    ).count()
                elif org.org_level == 4:  # 부
                    employee_count = Employee.objects.filter(
                        headquarters2=org.org_name,
                        employment_status='재직'
                    ).count()
                elif org.org_level == 5:  # 팀
                    employee_count = Employee.objects.filter(
                        final_department=org.org_name,
                        employment_status='재직'
                    ).count()
            
            node = {
                'id': org.id,
                'name': org.org_name,
                'code': org.org_code,
                'level': org.org_level,
                'level_name': org.get_org_level_display() if hasattr(org, 'get_org_level_display') else str(org.org_level),
                'position': leader_info['position'] if leader_info else '',
                'department': org.full_path,
                'leader': leader_info,
                'employee_count': employee_count,
                'description': org.description,
                'status': org.status,
                'children': build_org_tree(org.id, processed.copy())
            }
            nodes.append(node)
        
        return nodes
    
    try:
        org_data = build_org_tree()
        return JsonResponse({
            'success': True,
            'data': org_data,
            'source': 'OrganizationStructure'  # 데이터 소스 표시
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def organization_data_api_legacy(request):
    """조직도 데이터 JSON API - 레거시 버전 (Employee 테이블 사용)"""
    def build_tree(manager_id=None, processed=None):
        if processed is None:
            processed = set()
        
        # 순환 참조 방지
        if manager_id in processed:
            return []
        
        processed.add(manager_id)
        
        # 해당 관리자의 직속 부하들 조회
        if manager_id is None:
            # 최상위 관리자들 (manager가 None인 직원들)
            employees = Employee.objects.filter(
                manager__isnull=True,
                employment_status='재직'
            ).order_by('department', 'growth_level')
        else:
            employees = Employee.objects.filter(
                manager_id=manager_id,
                employment_status='재직'
            ).order_by('department', 'growth_level')
        
        nodes = []
        for emp in employees:
            node = {
                'id': emp.id,
                'name': emp.name,
                'position': emp.new_position,
                'department': emp.department,
                'job_type': emp.job_type,
                'level': emp.growth_level,
                'email': emp.email,
                'phone': emp.phone,
                'profile_image': emp.profile_image.url if emp.profile_image else None,
                'children': build_tree(emp.id, processed.copy())
            }
            nodes.append(node)
        
        return nodes
    
    try:
        org_data = build_tree()
        return JsonResponse({
            'success': True,
            'data': org_data,
            'source': 'Employee'  # 데이터 소스 표시
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# HR Dashboard View
def hr_dashboard_view(request):
    """HR 인력현황 대시보드 뷰"""
    return render(request, 'hr/dashboard.html')


def outsourced_dashboard_view(request):
    """외주인력 현황관리 대시보드 뷰"""
    return render(request, 'hr/outsourced_dashboard.html')


def workforce_dashboard_view(request):
    """주간 인력현황 대시보드"""
    return render(request, 'hr/workforce_dashboard.html')


def monthly_workforce_view(request):
    """월간 인력현황 대시보드"""
    return render(request, 'hr/monthly_workforce.html')


def full_workforce_view(request):
    """전체 인력현황 대시보드"""
    return render(request, 'hr/full_workforce.html')


def overseas_workforce_view(request):
    """월간 해외인력현황 뷰"""
    return render(request, 'hr/overseas_workforce.html')


# Organization Input Views
def organization_input_view(request):
    """조직 정보 입력 페이지"""
    return render(request, 'employees/organization_input.html')


def save_employee_api(request):
    """개별 직원 정보 저장 API"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            
            # Create or update employee
            employee = Employee(
                name=data.get('name'),
                email=data.get('email'),
                phone=data.get('phone', ''),
                company=data.get('company'),
                headquarters1=data.get('headquarters1', ''),
                department=data.get('department', ''),
                current_position=data.get('position'),
                new_position=data.get('position'),
                responsibility=data.get('responsibility', ''),
                hire_date=data.get('hire_date') if data.get('hire_date') else date.today(),
                employment_status='재직'
            )
            
            # Set manager if provided
            manager_id = data.get('manager')
            if manager_id:
                try:
                    manager = Employee.objects.get(id=manager_id)
                    employee.manager = manager
                except Employee.DoesNotExist:
                    pass
            
            employee.save()
            
            return JsonResponse({
                'success': True,
                'message': '직원 정보가 저장되었습니다.',
                'employee_id': employee.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'message': '잘못된 요청입니다.'}, status=400)


def bulk_upload_api(request):
    """Excel 일괄 업로드 API"""
    if request.method == 'POST':
        try:
            import json
            body = json.loads(request.body)
            data = body.get('data', [])
            
            created_count = 0
            errors = []
            
            for row in data:
                try:
                    # Map Excel columns to model fields
                    employee = Employee(
                        name=row.get('이름', ''),
                        email=row.get('이메일', ''),
                        phone=row.get('전화번호', ''),
                        company=row.get('회사', ''),
                        headquarters1=row.get('본부', ''),
                        department=row.get('부서', ''),
                        current_position=row.get('직급', ''),
                        new_position=row.get('직급', ''),
                        responsibility=row.get('직책', ''),
                        job_type=row.get('팀', ''),
                        employment_status='재직'
                    )
                    
                    # Parse hire date
                    hire_date_str = row.get('입사일')
                    if hire_date_str:
                        from datetime import datetime
                        try:
                            employee.hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
                        except:
                            employee.hire_date = date.today()
                    else:
                        employee.hire_date = date.today()
                    
                    # Find and set manager
                    manager_name = row.get('직속상사')
                    if manager_name:
                        try:
                            manager = Employee.objects.filter(name=manager_name).first()
                            if manager:
                                employee.manager = manager
                        except:
                            pass
                    
                    employee.save()
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row.get('이름', 'Unknown')}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'count': created_count,
                'errors': errors,
                'message': f'{created_count}개의 직원 정보가 업로드되었습니다.'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'message': '잘못된 요청입니다.'}, status=400)


def get_managers_api(request):
    """관리자 목록 조회 API"""
    try:
        # Get employees who can be managers (usually higher positions)
        managers = Employee.objects.filter(
            employment_status='재직',
            current_position__in=['팀장', '부장', '본부장', '이사', '상무', '전무', '부사장', '사장']
        ).values('id', 'name', 'current_position', 'department')
        
        manager_list = []
        for manager in managers:
            manager_list.append({
                'id': manager['id'],
                'name': manager['name'],
                'position': manager['current_position'],
                'department': manager['department']
            })
        
        return JsonResponse(manager_list, safe=False)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Organization Structure Management Views
def organization_structure_view(request):
    """조직 구조 관리 페이지"""
    return render(request, 'employees/organization_structure.html')

def organization_structure_upload_view(request):
    """조직 구조 업로드 페이지 (새 버전)"""
    return render(request, 'employees/organization_structure_upload.html')


@csrf_exempt
def upload_organization_structure(request):
    """조직 구조 Excel 업로드 처리 - 개선된 버전"""
    import json
    from datetime import datetime
    
    # 디버깅 로그
    print(f"[DEBUG] Request method: {request.method}")
    print(f"[DEBUG] Content-Type: {request.META.get('CONTENT_TYPE', 'None')}")
    print(f"[DEBUG] Request body length: {len(request.body) if request.body else 0}")
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': 'POST 메소드만 허용됩니다'
        }, status=405)
    
    try:
        # request body 파싱
        if not request.body:
            return JsonResponse({
                'success': False,
                'message': '요청 데이터가 없습니다'
            }, status=400)
        
        try:
            body = json.loads(request.body)
            data = body.get('data', [])
        except json.JSONDecodeError as e:
            print(f"[ERROR] JSON 파싱 오류: {e}")
            return JsonResponse({
                'success': False,
                'message': f'JSON 파싱 오류: {str(e)}'
            }, status=400)
        
        if not data:
            return JsonResponse({
                'success': False,
                'message': '업로드할 데이터가 없습니다'
            }, status=400)
        
        print(f"[DEBUG] 데이터 개수: {len(data)}")
        if data:
            print(f"[DEBUG] 첫 번째 행 키: {list(data[0].keys())}")
            print(f"[DEBUG] 첫 번째 행 데이터: {data[0]}")
            
            # 필드명 매핑 시도
            for idx, row in enumerate(data[:3]):  # 처음 3개만 디버깅
                print(f"[DEBUG] 행 {idx+1}:")
                for key, value in row.items():
                    print(f"  - {key}: {value}")
        
        if data:
            print(f"[DEBUG] 첫 번째 행 키: {list(data[0].keys())}")
            print(f"[DEBUG] 첫 번째 행 데이터: {data[0]}")
            
            # 필드명 매핑 시도
            for idx, row in enumerate(data[:3]):  # 처음 3개만 디버깅
                print(f"[DEBUG] 행 {idx+1}:")
                for key, value in row.items():
                    print(f"  - {key}: {value}")
        
        
        # 모델 임포트
        try:
            from employees.models_organization import OrganizationStructure, OrganizationUploadHistory
        except ImportError as e:
            print(f"[ERROR] 모델 임포트 실패: {e}")
            return JsonResponse({
                'success': False,
                'message': '서버 설정 오류'
            }, status=500)
        
        created_count = 0
        updated_count = 0
        errors = []
        
        # 업로드 기록 생성 (선택사항)
        try:
            upload_history = OrganizationUploadHistory.objects.create(
                file_name='Excel Upload',
                total_rows=len(data),
                status='processing',
                uploaded_by=request.user if request.user.is_authenticated else None
            )
        except Exception as e:
            print(f"[WARNING] 업로드 기록 생성 실패: {e}")
            upload_history = None
        
        # 데이터 처리
        for idx, row in enumerate(data):
            try:
                # 필수 필드 확인 (A,B,C,D 컬럼 또는 한글 컬럼명 지원)
                org_code = (row.get('조직코드') or row.get('A', '')).strip() if (row.get('조직코드') or row.get('A')) else ''
                org_name = (row.get('조직명') or row.get('B', '')).strip() if (row.get('조직명') or row.get('B')) else ''
                
                if not org_code:
                    errors.append({
                        'row': idx + 1,
                        'error': '조직코드 없음',
                        'data': row
                    })
                    continue
                
                if not org_name:
                    errors.append({
                        'row': idx + 1,
                        'error': '조직명 없음',
                        'data': row
                    })
                    continue
                
                # 조직 레벨 처리 (C 컬럼 또는 한글 컬럼명 지원)
                try:
                    org_level = int(row.get('조직레벨') or row.get('C', 1))
                except (ValueError, TypeError):
                    org_level = 1
                
                # 조직 생성 또는 업데이트
                org, created = OrganizationStructure.objects.get_or_create(
                    org_code=org_code,
                    defaults={
                        'org_name': org_name,
                        'org_level': org_level,
                        'status': row.get('상태') or row.get('F', 'active'),
                        'sort_order': int(row.get('정렬순서') or row.get('G', 0)) if (row.get('정렬순서') or row.get('G')) else 0,
                        'description': row.get('설명') or row.get('H', ''),
                    }
                )
                
                if not created:
                    # 기존 조직 업데이트
                    org.org_name = org_name
                    org.org_level = org_level
                    if row.get('상태') or row.get('F'):
                        org.status = row.get('상태') or row.get('F')
                    if row.get('설명') or row.get('H'):
                        org.description = row.get('설명') or row.get('H')
                    org.save()
                    updated_count += 1
                else:
                    created_count += 1
                
                # 상위 조직 설정 (D 컬럼 또는 한글 컬럼명 지원)
                parent_code = (row.get('상위조직코드') or row.get('D', '')).strip() if (row.get('상위조직코드') or row.get('D')) else ''
                if parent_code:
                    try:
                        parent = OrganizationStructure.objects.get(org_code=parent_code)
                        org.parent = parent
                        org.save()
                    except OrganizationStructure.DoesNotExist:
                        print(f"[WARNING] 상위조직 {parent_code} 찾을 수 없음")
                
            except Exception as e:
                print(f"[ERROR] 행 {idx + 1} 처리 오류: {e}")
                errors.append({
                    'row': idx + 1,
                    'error': str(e),
                    'data': row
                })
        
        # 업로드 기록 업데이트
        if upload_history:
            try:
                upload_history.status = 'completed'
                upload_history.success_count = created_count + updated_count
                upload_history.error_count = len(errors)
                upload_history.save()
            except:
                pass
        
        # 결과 반환
        result = {
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'total': created_count + updated_count,
            'errors': len(errors),
            'message': f'업로드 완료: {created_count}개 생성, {updated_count}개 업데이트'
        }
        
        if errors:
            result['error_details'] = errors[:10]  # 최대 10개 오류만 반환
        
        print(f"[SUCCESS] 업로드 결과: {result['message']}")
        return JsonResponse(result)
        
    except Exception as e:
        print(f"[CRITICAL] 업로드 실패: {e}")
        import traceback
        traceback.print_exc()
        
        return JsonResponse({
            'success': False,
            'message': f'서버 오류: {str(e)}'
        }, status=500)

def get_organization_tree(request):
    """조직 트리 구조 조회 - 안전 버전"""
    try:
        # 테이블 존재 확인
        try:
            from employees.models_organization import OrganizationStructure
            
            # 테이블 존재 테스트
            OrganizationStructure.objects.count()
        except Exception as e:
            print(f"OrganizationStructure 테이블 오류: {e}")
            return JsonResponse({
                'success': True,
                'tree': []
            })
        
        def build_tree(parent=None, depth=0):
            # 순환 참조 방지
            if depth > 10:
                return []
            
            try:
                # status 필드가 없을 수 있으므로 안전하게 처리
                orgs = OrganizationStructure.objects.filter(parent=parent)
                
                # status 필드 확인
                try:
                    orgs = orgs.filter(status='active')
                except:
                    pass  # status 필드가 없으면 모든 조직 포함
                
                # sort_order 필드 확인
                try:
                    orgs = orgs.order_by('sort_order', 'org_code')
                except:
                    orgs = orgs.order_by('org_code')
                
                tree = []
                for org in orgs:
                    node = {
                        'id': org.id,
                        'code': getattr(org, 'org_code', ''),
                        'name': getattr(org, 'org_name', ''),
                        'level': getattr(org, 'org_level', 1),
                        'employee_count': 0,  # 안전한 기본값
                        'children': build_tree(org, depth + 1)
                    }
                    
                    # employee_count 메소드가 있으면 사용
                    if hasattr(org, 'get_employee_count'):
                        try:
                            node['employee_count'] = org.get_employee_count()
                        except:
                            pass
                    
                    tree.append(node)
                
                return tree
                
            except Exception as e:
                print(f"build_tree 오류: {e}")
                return []
        
        tree = build_tree()
        
        return JsonResponse({
            'success': True,
            'tree': tree
        })
        
    except Exception as e:
        print(f"get_organization_tree 오류: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'tree': []
        }, status=200)  # 500 대신 200으로 반환


def get_organization_stats(request):
    """조직 통계 조회 - 안전 버전"""
    from datetime import datetime
    
    result = {
        'total_orgs': 0,
        'active_orgs': 0,
        'total_employees': 0,
        'last_update': '-'
    }
    
    try:
        # OrganizationStructure 쿼리
        try:
            result['total_orgs'] = OrganizationStructure.objects.count()
        except Exception as e:
            print(f"total_orgs error: {e}")
        
        try:
            result['active_orgs'] = OrganizationStructure.objects.filter(status='active').count()
        except Exception as e:
            # status 필드가 없을 수 있음
            result['active_orgs'] = result['total_orgs']
            print(f"active_orgs error: {e}")
        
        # Employee 쿼리
        try:
            result['total_employees'] = Employee.objects.filter(employment_status='재직').count()
        except Exception as e:
            # employment_status 필드가 없을 수 있음
            try:
                result['total_employees'] = Employee.objects.count()
            except:
                result['total_employees'] = 0
            print(f"total_employees error: {e}")
        
        # 최종 업데이트 시간
        try:
            last_org = OrganizationStructure.objects.order_by('-updated_at').first()
            if last_org and hasattr(last_org, 'updated_at'):
                result['last_update'] = last_org.updated_at.strftime('%Y-%m-%d')
        except Exception as e:
            # updated_at 필드가 없을 수 있음
            result['last_update'] = datetime.now().strftime('%Y-%m-%d')
            print(f"last_update error: {e}")
        
        return JsonResponse(result)
        
    except Exception as e:
        # 최악의 경우에도 기본값 반환
        print(f"Critical error in get_organization_stats: {e}")
        return JsonResponse(result)

@csrf_exempt
def delete_organization_data(request):
    """조직 데이터 삭제 API"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST 요청만 허용됩니다.'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        delete_type = data.get('type', 'all')  # 'all' 또는 'specific'
        org_code = data.get('org_code', '')
        
        from employees.models_organization import OrganizationStructure, OrganizationUploadHistory, EmployeeOrganizationMapping
        
        if delete_type == 'all':
            # 모든 조직 데이터 삭제
            try:
                mapping_count = EmployeeOrganizationMapping.objects.count()
                history_count = OrganizationUploadHistory.objects.count()
                org_count = OrganizationStructure.objects.count()
                
                EmployeeOrganizationMapping.objects.all().delete()
                OrganizationUploadHistory.objects.all().delete()
                OrganizationStructure.objects.all().delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'모든 조직 데이터 삭제 완료 (조직: {org_count}개, 매핑: {mapping_count}개, 히스토리: {history_count}개)'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'삭제 중 오류 발생: {str(e)}'
                }, status=500)
            
        elif delete_type == 'specific' and org_code:
            # 특정 조직 및 하위 조직 삭제
            try:
                org = OrganizationStructure.objects.get(org_code=org_code)
                descendants = org.get_descendants(include_self=True)
                count = len(descendants)
                
                # 하위부터 삭제 (외래키 제약조건)
                for desc in reversed(descendants):
                    desc.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'조직 {org_code} 및 하위 {count}개 조직 삭제 완료'
                })
                
            except OrganizationStructure.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': f'조직코드 {org_code}를 찾을 수 없습니다.'
                }, status=404)
        
        else:
            return JsonResponse({
                'success': False,
                'message': '잘못된 삭제 타입입니다.'
            }, status=400)
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'삭제 중 오류 발생: {str(e)}'
        }, status=500)

@csrf_exempt
def save_organization(request):
    """개별 조직 저장"""
    if request.method == 'POST':
        try:
            import json
            # Using models imported at top of file
            
            data = json.loads(request.body)
            
            org = OrganizationStructure(
                org_code=data.get('org_code'),
                org_name=data.get('org_name'),
                org_level=int(data.get('org_level')),
                status='active'
            )
            
            # Set parent if provided
            parent_id = data.get('parent_org')
            if parent_id:
                try:
                    parent = OrganizationStructure.objects.get(id=parent_id)
                    org.parent = parent
                except:
                    pass
            
            org.save()
            
            return JsonResponse({
                'success': True,
                'message': '조직이 저장되었습니다.'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'message': '잘못된 요청'}, status=400)


def download_org_sample(request):
    """조직 구조 샘플 데이터 다운로드"""
    import io
    from django.http import HttpResponse
    
    try:
        import xlsxwriter
        
        # Create Excel file in memory
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('조직구조')
        
    except ImportError:
        # xlsxwriter가 없는 경우 openpyxl 사용
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = '조직구조'
            
        except ImportError:
            # 둘 다 없는 경우 CSV로 대체
            import csv
            from django.http import HttpResponse
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename=조직구조_샘플데이터.csv'
            
            writer = csv.writer(response)
            headers = ['조직코드', '조직명', '조직레벨', '상위조직코드', '조직장', '상태', '정렬순서', '설명']
            writer.writerow(headers)
            
            # Sample data
            sample_data = [
                ['GRP001', 'OK금융그룹', 1, '', '', 'active', 1, 'OK금융그룹 지주회사'],
                ['COM001', 'OK저축은행', 2, 'GRP001', '', 'active', 1, '저축은행'],
                ['COM002', 'OK캐피탈', 2, 'GRP001', '', 'active', 2, '캐피탈'],
                ['HQ001', '경영지원본부', 3, 'COM001', '', 'active', 1, ''],
                ['HQ003', '디지털본부', 3, 'COM001', '', 'active', 3, ''],
                ['DEPT001', 'IT개발부', 4, 'HQ003', '', 'active', 1, ''],
                ['TEAM001', '개발1팀', 5, 'DEPT001', '', 'active', 1, '코어뱅킹 시스템'],
            ]
            
            for row in sample_data:
                writer.writerow(row)
                
            return response
    
    # Define headers
    headers = ['조직코드', '조직명', '조직레벨', '상위조직코드', '조직장', '상태', '정렬순서', '설명']
    
    # Sample data - OK Financial Group structure
    sample_data = [
        ['GRP001', 'OK금융그룹', 1, '', '', 'active', 1, 'OK금융그룹 지주회사'],
        
        # 계열사 (Level 2)
        ['COM001', 'OK저축은행', 2, 'GRP001', '', 'active', 1, '저축은행'],
        ['COM002', 'OK캐피탈', 2, 'GRP001', '', 'active', 2, '캐피탈'],
        ['COM003', 'OK손해보험', 2, 'GRP001', '', 'active', 3, '손해보험'],
        
        # 본부 (Level 3) - OK저축은행
        ['HQ001', '경영지원본부', 3, 'COM001', '', 'active', 1, ''],
        ['HQ002', '영업본부', 3, 'COM001', '', 'active', 2, ''],
        ['HQ003', '디지털본부', 3, 'COM001', '', 'active', 3, ''],
        ['HQ004', '리스크관리본부', 3, 'COM001', '', 'active', 4, ''],
        
        # 부 (Level 4) - 디지털본부
        ['DEPT001', 'IT개발부', 4, 'HQ003', '', 'active', 1, ''],
        ['DEPT002', 'IT운영부', 4, 'HQ003', '', 'active', 2, ''],
        ['DEPT003', '디지털마케팅부', 4, 'HQ003', '', 'active', 3, ''],
        
        # 팀 (Level 5) - IT개발부
        ['TEAM001', '개발1팀', 5, 'DEPT001', '', 'active', 1, '코어뱅킹 시스템'],
        ['TEAM002', '개발2팀', 5, 'DEPT001', '', 'active', 2, '모바일/웹 개발'],
        ['TEAM003', 'AI개발팀', 5, 'DEPT001', '', 'active', 3, 'AI/ML 솔루션'],
    ]
    
    try:
        # xlsxwriter를 사용하는 경우
        if 'xlsxwriter' in locals():
            # Write headers
            for col, header in enumerate(headers):
                worksheet.write(0, col, header)
            
            # Write sample data
            for row_num, row_data in enumerate(sample_data, 1):
                for col, value in enumerate(row_data):
                    worksheet.write(row_num, col, value)
            
            workbook.close()
            
            # Prepare response
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=조직구조_샘플데이터.xlsx'
            
            return response
            
        # openpyxl을 사용하는 경우
        elif 'openpyxl' in locals():
            # Write headers
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Write sample data
            for row_num, row_data in enumerate(sample_data, 2):
                for col, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col, value=value)
            
            # Save to BytesIO
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=조직구조_샘플데이터.xlsx'
            
            return response
    
    except Exception as e:
        # 모든 Excel 라이브러리가 실패한 경우 CSV로 대체
        import csv
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename=조직구조_샘플데이터.csv'
        
        # BOM 추가 (Excel에서 한글 제대로 표시)
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for row in sample_data:
            writer.writerow(row)
            
        return response
