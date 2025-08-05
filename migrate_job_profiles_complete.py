#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
OK금융그룹 직무기술서 PDF 완전 파싱 - 37개 전체 직무
"""

import os
import sys
import io
import json
import re
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd
from tqdm import tqdm
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 한글 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 로깅 설정
def setup_logging(output_dir: str):
    """로깅 설정"""
    log_dir = os.path.join(output_dir, 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    log_file = os.path.join(log_dir, f'migration_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    return logging.getLogger(__name__)


class CompleteJobProfileParser:
    """OK금융그룹 전체 직무기술서 파서"""
    
    def __init__(self, pdf_path: str, output_dir: str):
        self.pdf_path = pdf_path
        self.output_dir = output_dir
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # 전체 직무 목록 (첫 페이지 기준)
        self.all_jobs = {
            'Non-PL': {
                'IT기획': ['시스템기획'],
                'IT개발': ['시스템개발'],
                'IT운영': ['시스템관리', '서비스운영'],
                '경영관리': ['감사', 'HRM', 'HRD', '경영지원', '비서', 'PR', '경영기획', 
                            '디자인', '리스크관리', '마케팅', '스포츠사무관리', '자금', 
                            '재무회계', '정보보안', '준법지원', '총무'],
                '투자금융': ['IB금융'],
                '기업금융': ['기업영업기획', '기업여신심사', '기업여신관리'],
                '기업영업': ['여신영업'],
                '리테일금융': ['데이터/통계', '플랫폼/핀테크', 'NPL영업기획', '리테일심사기획', 
                              'PL기획', '모기지기획', '수신기획', '수신영업']
            },
            'PL': {
                '고객지원': ['여신고객지원', '사무지원', '수신고객지원', '채권관리지원']
            }
        }
        
        # 직군 카테고리 매핑
        self.category_mapping = {
            'IT기획': 'IT/디지털',
            'IT개발': 'IT/디지털', 
            'IT운영': 'IT/디지털',
            '경영관리': '경영지원',
            '투자금융': '금융',
            '기업금융': '금융',
            '기업영업': '영업',
            '리테일금융': '금융',
            '고객지원': '고객서비스'
        }
        
        self.parsed_jobs = []
    
    def extract_text_from_pdf(self) -> List[Tuple[int, str]]:
        """PDF에서 텍스트 추출"""
        self.logger.info(f"PDF 파일 읽기 시작: {self.pdf_path}")
        
        try:
            doc = fitz.open(self.pdf_path)
            pages_text = []
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text()
                pages_text.append((page_num + 1, text))
                
            doc.close()
            self.logger.info(f"총 {len(pages_text)}페이지 추출 완료")
            return pages_text
            
        except Exception as e:
            self.logger.error(f"PDF 읽기 오류: {str(e)}")
            raise
    
    def parse_job_profile_pages(self, pages_text: List[Tuple[int, str]]) -> List[Dict]:
        """Job Profile 페이지 파싱"""
        jobs = []
        
        for i in range(len(pages_text)):
            page_num, text = pages_text[i]
            
            # Job Profile 헤더 페이지 찾기
            if 'Job Profile' in text and '직군' in text:
                job_data = {'page_number': page_num}
                
                # 직군 구분 (Non-PL / PL)
                if 'Non-PL직군' in text:
                    job_data['group'] = 'Non-PL'
                elif 'PL직군' in text:
                    job_data['group'] = 'PL'
                else:
                    continue
                
                # 직종 추출
                # 패턴: "직군경영관리직종감사" 또는 "직군IT기획직종시스템기획"
                type_match = re.search(r'직군([가-힣/]+)직종([가-힣/\s]+)', text)
                if type_match:
                    job_type = type_match.group(1).strip()
                    job_name_hint = type_match.group(2).strip()
                    
                    job_data['job_type'] = job_type
                    job_data['job_category'] = self.category_mapping.get(job_type, '기타')
                    
                    self.logger.info(f"페이지 {page_num}: {job_data['group']} - {job_type}")
                
                # 다음 페이지에서 직무 상세 정보 추출
                if i + 1 < len(pages_text):
                    next_page_num, next_text = pages_text[i + 1]
                    
                    # 직무명 추출
                    job_match = re.search(r'직무:\s*([^\n]+)', next_text)
                    if job_match:
                        job_data['job_title'] = job_match.group(1).strip()
                        
                        # 역할 및 책임 추출
                        job_data['responsibilities'] = self.extract_responsibilities(next_text)
                        
                        # 그 다음 페이지에서 기술/지식 추출
                        if i + 2 < len(pages_text):
                            skill_page_num, skill_text = pages_text[i + 2]
                            job_data['basic_skills'] = self.extract_skills(skill_text, '기본기술&지식')
                            job_data['advanced_skills'] = self.extract_skills(skill_text, '응용기술&지식')
                        
                        jobs.append(job_data)
                        self.logger.info(f"  -> 직무: {job_data['job_title']}")
        
        return jobs
    
    def extract_responsibilities(self, text: str) -> List[str]:
        """핵심역할&책임 추출"""
        responsibilities = []
        
        # 핵심역할&책임 섹션 찾기
        if '핵심역할&책임' in text:
            lines = text.split('\n')
            in_section = False
            current_item = []
            
            for line in lines:
                line = line.strip()
                
                if '핵심역할&책임' in line:
                    in_section = True
                    continue
                
                if in_section:
                    # 다음 섹션 시작 시 종료
                    if any(keyword in line for keyword in ['기본기술&지식', '응용기술&지식', '기술&지식']):
                        break
                    
                    # 새로운 항목 시작 (짧은 제목)
                    if line and len(line.split()) <= 4 and not any(c.isdigit() for c in line) and '정의' not in line:
                        if current_item:
                            responsibilities.append(' '.join(current_item))
                        current_item = [line + ':']
                    elif line and current_item:
                        # 설명 추가
                        current_item.append(line)
            
            # 마지막 항목 추가
            if current_item:
                responsibilities.append(' '.join(current_item))
        
        return responsibilities
    
    def extract_skills(self, text: str, section_name: str) -> List[str]:
        """기술&지식 추출"""
        skills = []
        
        if section_name in text:
            start_idx = text.find(section_name)
            # 다음 섹션까지 또는 끝까지
            if section_name == '기본기술&지식' and '응용기술&지식' in text:
                end_idx = text.find('응용기술&지식', start_idx)
            else:
                end_idx = len(text)
            
            section_text = text[start_idx:end_idx]
            lines = section_text.split('\n')
            
            current_skill = []
            for line in lines[1:]:  # 헤더 제외
                line = line.strip()
                
                if not line or '정의' in line:
                    continue
                
                # 새로운 스킬 항목 (짧은 제목)
                if len(line.split()) <= 4 and not any(c.isdigit() for c in line):
                    if current_skill:
                        skills.append(' '.join(current_skill))
                    current_skill = [line + ':']
                elif line and current_skill:
                    current_skill.append(line)
            
            # 마지막 스킬 추가
            if current_skill:
                skills.append(' '.join(current_skill))
        
        return skills
    
    def verify_completeness(self, parsed_jobs: List[Dict]) -> Dict[str, List[str]]:
        """파싱 완전성 검증"""
        found_jobs = {}
        missing_jobs = {}
        
        # 파싱된 직무 정리
        for job in parsed_jobs:
            group = job.get('group', 'Unknown')
            job_type = job.get('job_type', 'Unknown')
            job_title = job.get('job_title', 'Unknown')
            
            if group not in found_jobs:
                found_jobs[group] = {}
            if job_type not in found_jobs[group]:
                found_jobs[group][job_type] = []
            found_jobs[group][job_type].append(job_title)
        
        # 누락된 직무 확인
        for group, types in self.all_jobs.items():
            if group not in missing_jobs:
                missing_jobs[group] = {}
            
            for job_type, jobs in types.items():
                for job in jobs:
                    found = False
                    if group in found_jobs and job_type in found_jobs[group]:
                        if any(job in found_job for found_job in found_jobs[group][job_type]):
                            found = True
                    
                    if not found:
                        if job_type not in missing_jobs[group]:
                            missing_jobs[group][job_type] = []
                        missing_jobs[group][job_type].append(job)
        
        return found_jobs, missing_jobs
    
    def parse_all(self) -> List[Dict]:
        """전체 파싱 실행"""
        self.logger.info("OK금융그룹 직무기술서 전체 파싱 시작")
        
        # PDF 텍스트 추출
        pages_text = self.extract_text_from_pdf()
        
        # 직무 파싱
        self.parsed_jobs = self.parse_job_profile_pages(pages_text)
        
        # 완전성 검증
        found_jobs, missing_jobs = self.verify_completeness(self.parsed_jobs)
        
        # 결과 출력
        self.logger.info(f"\n총 파싱된 직무: {len(self.parsed_jobs)}개")
        
        # 그룹별 집계
        for group in ['Non-PL', 'PL']:
            group_jobs = [j for j in self.parsed_jobs if j.get('group') == group]
            self.logger.info(f"{group} 직군: {len(group_jobs)}개")
        
        # 누락된 직무 확인
        total_missing = 0
        for group, types in missing_jobs.items():
            for job_type, jobs in types.items():
                if jobs:
                    total_missing += len(jobs)
                    self.logger.warning(f"누락: {group} > {job_type} > {jobs}")
        
        if total_missing > 0:
            self.logger.warning(f"총 누락된 직무: {total_missing}개")
        else:
            self.logger.info("✅ 모든 직무가 성공적으로 파싱되었습니다!")
        
        return self.parsed_jobs


class JobDataExporter:
    """직무 데이터 내보내기"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def export_to_excel(self, jobs_data: List[Dict]) -> str:
        """Excel 파일로 내보내기"""
        excel_dir = os.path.join(self.output_dir, 'excel')
        os.makedirs(excel_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = os.path.join(excel_dir, f'OK금융그룹_직무기술서_전체_{timestamp}.xlsx')
        
        # 데이터프레임 생성
        df_data = []
        for job in jobs_data:
            df_data.append({
                '구분': job.get('group', ''),
                '직종': job.get('job_type', ''),
                '직무명': job.get('job_title', ''),
                '직군': job.get('job_category', ''),
                '페이지': job.get('page_number', ''),
                '핵심역할 및 책임': '\n'.join(job.get('responsibilities', [])),
                '필수역량': '\n'.join(job.get('basic_skills', [])),
                '우대역량': '\n'.join(job.get('advanced_skills', []))
            })
        
        df = pd.DataFrame(df_data)
        
        # Excel 저장
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 전체 시트
            df.to_excel(writer, sheet_name='전체직무(37개)', index=False)
            
            # Non-PL / PL 구분 시트
            for group in ['Non-PL', 'PL']:
                group_df = df[df['구분'] == group]
                if not group_df.empty:
                    sheet_name = f'{group}직군({len(group_df)}개)'
                    group_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 직종별 시트
            for job_type in df['직종'].unique():
                if pd.notna(job_type):
                    type_df = df[df['직종'] == job_type]
                    sheet_name = f'{job_type}({len(type_df)}개)'[:31]
                    type_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 스타일 적용
        wb = openpyxl.load_workbook(excel_file)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 헤더 스타일
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 열 너비 조정
            column_widths = {
                'A': 10,   # 구분
                'B': 15,   # 직종
                'C': 25,   # 직무명
                'D': 15,   # 직군
                'E': 10,   # 페이지
                'F': 60,   # 역할
                'G': 50,   # 필수역량
                'H': 50    # 우대역량
            }
            
            for col, width in column_widths.items():
                if col in ws.column_dimensions:
                    ws.column_dimensions[col].width = width
            
            # 텍스트 줄바꿈
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if cell.column in [6, 7, 8]:  # F, G, H 열
                        ws.row_dimensions[cell.row].height = 100
        
        wb.save(excel_file)
        self.logger.info(f"Excel 내보내기 완료: {excel_file}")
        return excel_file
    
    def export_to_json(self, jobs_data: List[Dict]) -> str:
        """JSON 파일로 내보내기"""
        json_dir = os.path.join(self.output_dir, 'json')
        os.makedirs(json_dir, exist_ok=True)
        
        json_file = os.path.join(json_dir, 'ok_job_profiles_complete.json')
        
        # 구조화된 형태로 저장
        structured_data = {
            'metadata': {
                'total_jobs': len(jobs_data),
                'non_pl_jobs': len([j for j in jobs_data if j.get('group') == 'Non-PL']),
                'pl_jobs': len([j for j in jobs_data if j.get('group') == 'PL']),
                'exported_at': datetime.now().isoformat()
            },
            'jobs': jobs_data
        }
        
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(structured_data, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"JSON 내보내기 완료: {json_file}")
        return json_file
    
    def generate_summary_report(self, jobs_data: List[Dict]) -> str:
        """요약 리포트 생성"""
        report_file = os.path.join(self.output_dir, 'COMPLETE_SUMMARY.md')
        
        # 통계 계산
        non_pl_jobs = [j for j in jobs_data if j.get('group') == 'Non-PL']
        pl_jobs = [j for j in jobs_data if j.get('group') == 'PL']
        
        # 직종별 집계
        job_type_stats = {}
        for job in jobs_data:
            job_type = job.get('job_type', '미분류')
            if job_type not in job_type_stats:
                job_type_stats[job_type] = []
            job_type_stats[job_type].append(job.get('job_title', ''))
        
        # 리포트 작성
        report = f"""# OK금융그룹 직무기술서 전체 마이그레이션 완료

## 📊 최종 결과
- **추출 일시**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- **총 직무 수**: {len(jobs_data)}개
  - Non-PL 직군: {len(non_pl_jobs)}개
  - PL 직군: {len(pl_jobs)}개

## 📋 직종별 상세 현황

### Non-PL 직군 (8개 직종, 33개 직무)
"""
        
        # Non-PL 직종별 출력
        for job_type, jobs in sorted(job_type_stats.items()):
            job_list = [j for j in jobs_data if j.get('job_type') == job_type and j.get('group') == 'Non-PL']
            if job_list:
                report += f"\n#### {job_type} ({len(job_list)}개)\n"
                for job in job_list:
                    report += f"- {job['job_title']}\n"
        
        report += "\n### PL 직군 (1개 직종, 4개 직무)\n"
        
        # PL 직종별 출력
        for job_type, jobs in sorted(job_type_stats.items()):
            job_list = [j for j in jobs_data if j.get('job_type') == job_type and j.get('group') == 'PL']
            if job_list:
                report += f"\n#### {job_type} ({len(job_list)}개)\n"
                for job in job_list:
                    report += f"- {job['job_title']}\n"
        
        report += f"""
## ✅ 검증 결과
- PDF 명시: Non-PL 33개 + PL 4개 = 총 37개
- 실제 추출: {len(jobs_data)}개
- **상태**: {'✅ 완전 일치' if len(jobs_data) == 37 else '⚠️ 불일치'}

## 📁 생성된 파일
1. Excel: `excel/OK금융그룹_직무기술서_전체_*.xlsx`
2. JSON: `json/ok_job_profiles_complete.json`
3. 로그: `logs/migration_*.log`

---
*OK Job Profile Complete Migration Tool*
"""
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report)
        
        self.logger.info(f"요약 리포트 생성: {report_file}")
        return report_file


def main():
    """메인 실행 함수"""
    pdf_path = r"C:/Users/apro/OneDrive/Desktop/설명회자료/OK_Job Profile.pdf"
    output_dir = r"C:/Users/apro/OneDrive/Desktop/설명회자료/job_profile_complete"
    
    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    # 로깅 설정
    logger = setup_logging(output_dir)
    logger.info("="*60)
    logger.info("OK금융그룹 직무기술서 완전 마이그레이션 시작")
    logger.info("목표: Non-PL 33개 + PL 4개 = 총 37개 직무")
    logger.info("="*60)
    
    try:
        # 1. PDF 파싱
        parser = CompleteJobProfileParser(pdf_path, output_dir)
        jobs_data = parser.parse_all()
        
        # 2. 데이터 내보내기
        exporter = JobDataExporter(output_dir)
        excel_file = exporter.export_to_excel(jobs_data)
        json_file = exporter.export_to_json(jobs_data)
        report_file = exporter.generate_summary_report(jobs_data)
        
        # 완료
        logger.info("="*60)
        logger.info("마이그레이션 완료!")
        logger.info("="*60)
        
        print(f"\n✅ OK금융그룹 직무기술서 전체 마이그레이션 완료!")
        print(f"📊 추출 결과: {len(jobs_data)}개 직무 / 목표 37개")
        print(f"📁 출력 위치: {output_dir}")
        print(f"📄 Excel: {os.path.basename(excel_file)}")
        print(f"💾 JSON: {os.path.basename(json_file)}")
        print(f"📋 리포트: {os.path.basename(report_file)}")
        
    except Exception as e:
        logger.error(f"오류 발생: {str(e)}", exc_info=True)
        print(f"\n❌ 오류 발생: {str(e)}")


if __name__ == '__main__':
    main()