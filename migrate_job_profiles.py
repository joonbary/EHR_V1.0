#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
OK금융그룹 직무기술서 PDF 파싱 및 데이터베이스 마이그레이션 자동화 스크립트

이 스크립트는 다음 작업을 수행합니다:
1. PDF에서 직무별 섹션 자동 감지 및 분리
2. 직무 정보 구조화 (직군/직종/직무/역할/자격요건)
3. OK금융그룹 직무분류체계 매핑
4. 데이터 검증 및 정제
5. Django 데이터베이스 마이그레이션
6. JSON/Excel 백업 생성
"""

import os
import sys
import io
import json
import re
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd
from tqdm import tqdm
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 한글 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 로깅 설정
def setup_logging(output_dir: str):
    """로깅 설정"""
    log_dir = os.path.join(output_dir, 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    log_file = os.path.join(log_dir, f'migration_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    return logging.getLogger(__name__)


class JobProfilePDFParser:
    """OK금융그룹 직무기술서 PDF 파서"""
    
    def __init__(self, pdf_path: str, output_dir: str):
        self.pdf_path = pdf_path
        self.output_dir = output_dir
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # OK금융그룹 직무분류체계
        self.job_categories = {
            'IT/디지털': {
                'id': 'it_digital',
                'job_types': {
                    '시스템개발': ['백엔드개발', '프론트엔드개발', '풀스택개발', '모바일개발'],
                    'IT기획': ['IT전략기획', '프로젝트관리', '비즈니스분석'],
                    '데이터': ['데이터분석', '데이터엔지니어링', 'AI/ML엔지니어'],
                    '정보보안': ['보안관제', '보안컨설팅', '침해대응'],
                    'IT인프라': ['시스템엔지니어', '네트워크엔지니어', '클라우드엔지니어']
                }
            },
            '경영지원': {
                'id': 'management_support',
                'job_types': {
                    '인사': ['인사기획', '인재개발', '조직문화', '노무관리'],
                    '재무': ['재무기획', '자금관리', '회계', '세무'],
                    '총무': ['총무기획', '구매관리', '자산관리'],
                    '법무': ['법무지원', '컴플라이언스', '계약관리'],
                    '홍보': ['대외홍보', '브랜드관리', '사내커뮤니케이션']
                }
            },
            '영업/마케팅': {
                'id': 'sales_marketing',
                'job_types': {
                    '개인영업': ['개인금융영업', 'PB영업', '디지털영업'],
                    '기업영업': ['기업금융영업', 'IB영업', '글로벌영업'],
                    '마케팅': ['마케팅전략', '디지털마케팅', '상품기획'],
                    '고객관리': ['고객서비스', 'CS기획', 'VOC관리']
                }
            },
            '리스크/심사': {
                'id': 'risk_audit',
                'job_types': {
                    '리스크관리': ['신용리스크', '시장리스크', '운영리스크'],
                    '여신심사': ['개인여신심사', '기업여신심사', '심사기획'],
                    '감사': ['내부감사', '준법감시', '리스크모니터링']
                }
            }
        }
        
        # 직무 정보 패턴
        self.patterns = {
            'job_title': re.compile(r'^(?:직무명|직무)\s*[:：]\s*(.+)$', re.MULTILINE),
            'job_category': re.compile(r'^(?:직군|분야)\s*[:：]\s*(.+)$', re.MULTILINE),
            'job_type': re.compile(r'^(?:직종|직무군)\s*[:：]\s*(.+)$', re.MULTILINE),
            'responsibility': re.compile(r'^(?:주요\s*업무|담당업무|역할)\s*[:：]?\s*\n((?:[-•·]\s*.+\n?)+)', re.MULTILINE),
            'qualification': re.compile(r'^(?:자격요건|필요역량|요구사항)\s*[:：]?\s*\n((?:[-•·]\s*.+\n?)+)', re.MULTILINE),
            'preferred': re.compile(r'^(?:우대사항|우대조건)\s*[:：]?\s*\n((?:[-•·]\s*.+\n?)+)', re.MULTILINE),
            'section_divider': re.compile(r'^[-=]{3,}|^[━─]{3,}', re.MULTILINE)
        }
        
        self.parsed_jobs = []
        self.validation_errors = []
    
    def extract_text_from_pdf(self) -> List[Tuple[int, str]]:
        """PDF에서 텍스트 추출"""
        self.logger.info(f"PDF 파일 읽기 시작: {self.pdf_path}")
        
        try:
            doc = fitz.open(self.pdf_path)
            pages_text = []
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text()
                pages_text.append((page_num + 1, text))
                
            doc.close()
            self.logger.info(f"총 {len(pages_text)}페이지 추출 완료")
            return pages_text
            
        except Exception as e:
            self.logger.error(f"PDF 읽기 오류: {str(e)}")
            raise
    
    def detect_job_sections(self, text: str) -> List[str]:
        """텍스트에서 직무별 섹션 감지 및 분리"""
        # 섹션 구분자로 분리
        sections = self.patterns['section_divider'].split(text)
        
        job_sections = []
        for section in sections:
            # 직무명이 포함된 섹션만 추출
            if self.patterns['job_title'].search(section):
                job_sections.append(section.strip())
        
        self.logger.info(f"{len(job_sections)}개의 직무 섹션 감지됨")
        return job_sections
    
    def parse_job_section(self, section: str, page_num: int) -> Optional[Dict]:
        """개별 직무 섹션 파싱"""
        job_data = {
            'page_number': page_num,
            'raw_text': section
        }
        
        # 직무명 추출
        job_title_match = self.patterns['job_title'].search(section)
        if job_title_match:
            job_data['job_title'] = job_title_match.group(1).strip()
        else:
            self.logger.warning(f"페이지 {page_num}: 직무명을 찾을 수 없음")
            return None
        
        # 직군/직종 추출
        category_match = self.patterns['job_category'].search(section)
        if category_match:
            job_data['job_category'] = category_match.group(1).strip()
        
        type_match = self.patterns['job_type'].search(section)
        if type_match:
            job_data['job_type'] = type_match.group(1).strip()
        
        # 역할 및 책임 추출
        responsibility_match = self.patterns['responsibility'].search(section)
        if responsibility_match:
            responsibilities = responsibility_match.group(1).strip()
            job_data['responsibilities'] = self._parse_bullet_points(responsibilities)
        
        # 자격요건 추출
        qualification_match = self.patterns['qualification'].search(section)
        if qualification_match:
            qualifications = qualification_match.group(1).strip()
            job_data['qualifications'] = self._parse_bullet_points(qualifications)
        
        # 우대사항 추출
        preferred_match = self.patterns['preferred'].search(section)
        if preferred_match:
            preferred = preferred_match.group(1).strip()
            job_data['preferred'] = self._parse_bullet_points(preferred)
        
        # 직무분류체계 매핑
        job_data = self._map_to_category_system(job_data)
        
        return job_data
    
    def _parse_bullet_points(self, text: str) -> List[str]:
        """불릿 포인트 텍스트 파싱"""
        lines = text.strip().split('\n')
        points = []
        
        for line in lines:
            # 불릿 기호 제거 및 정리
            cleaned = re.sub(r'^[-•·]\s*', '', line).strip()
            if cleaned:
                points.append(cleaned)
        
        return points
    
    def _map_to_category_system(self, job_data: Dict) -> Dict:
        """OK금융그룹 직무분류체계에 매핑"""
        category_name = job_data.get('job_category', '')
        type_name = job_data.get('job_type', '')
        
        # 직군 매핑
        mapped_category = None
        mapped_type = None
        
        for cat_key, cat_value in self.job_categories.items():
            if cat_key in category_name or category_name in cat_key:
                mapped_category = cat_key
                job_data['category_id'] = cat_value['id']
                
                # 직종 매핑
                for type_key, jobs in cat_value['job_types'].items():
                    if type_key in type_name or type_name in type_key:
                        mapped_type = type_key
                        job_data['type_id'] = f"{cat_value['id']}_{type_key.replace('/', '_')}"
                        break
                break
        
        if not mapped_category:
            self.logger.warning(f"직군 매핑 실패: {category_name}")
            job_data['category_id'] = 'undefined'
        
        if not mapped_type:
            self.logger.warning(f"직종 매핑 실패: {type_name}")
            job_data['type_id'] = 'undefined'
        
        job_data['mapped_category'] = mapped_category
        job_data['mapped_type'] = mapped_type
        
        return job_data
    
    def validate_job_data(self, job_data: Dict) -> Tuple[bool, List[str]]:
        """직무 데이터 검증"""
        errors = []
        
        # 필수 필드 검증
        required_fields = ['job_title', 'responsibilities', 'qualifications']
        for field in required_fields:
            if field not in job_data or not job_data[field]:
                errors.append(f"필수 필드 누락: {field}")
        
        # 역할/자격요건 최소 개수 검증
        if 'responsibilities' in job_data and len(job_data['responsibilities']) < 3:
            errors.append("역할 및 책임이 3개 미만")
        
        if 'qualifications' in job_data and len(job_data['qualifications']) < 2:
            errors.append("자격요건이 2개 미만")
        
        # 직무분류 매핑 검증
        if job_data.get('category_id') == 'undefined':
            errors.append("직군 분류 매핑 실패")
        
        is_valid = len(errors) == 0
        return is_valid, errors
    
    def parse_all_jobs(self) -> List[Dict]:
        """전체 PDF 파싱 실행"""
        self.logger.info("직무기술서 파싱 시작")
        
        # PDF 텍스트 추출
        pages_text = self.extract_text_from_pdf()
        
        # 전체 텍스트 결합
        full_text = "\n\n".join([text for _, text in pages_text])
        
        # 직무 섹션 감지
        job_sections = self.detect_job_sections(full_text)
        
        # 각 섹션 파싱
        with tqdm(total=len(job_sections), desc="직무 파싱 진행") as pbar:
            for idx, section in enumerate(job_sections):
                # 해당 섹션이 속한 페이지 찾기
                page_num = 1
                for p_num, p_text in pages_text:
                    if section[:100] in p_text:
                        page_num = p_num
                        break
                
                job_data = self.parse_job_section(section, page_num)
                
                if job_data:
                    # 검증
                    is_valid, errors = self.validate_job_data(job_data)
                    job_data['is_valid'] = is_valid
                    job_data['validation_errors'] = errors
                    
                    if not is_valid:
                        self.logger.warning(f"검증 실패 - {job_data.get('job_title', 'Unknown')}: {errors}")
                    
                    self.parsed_jobs.append(job_data)
                
                pbar.update(1)
        
        self.logger.info(f"총 {len(self.parsed_jobs)}개 직무 파싱 완료")
        valid_count = sum(1 for job in self.parsed_jobs if job['is_valid'])
        self.logger.info(f"유효한 직무: {valid_count}개")
        
        return self.parsed_jobs


class JobProfileMigrator:
    """직무기술서 데이터베이스 마이그레이션"""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # Django 설정
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ehr_system.settings')
        import django
        django.setup()
    
    def migrate_to_database(self, jobs_data: List[Dict]) -> Dict[str, int]:
        """Django 데이터베이스로 마이그레이션"""
        from job_profiles.models import JobCategory, JobType, JobRole, JobProfile
        from django.db import transaction
        
        stats = {
            'categories_created': 0,
            'types_created': 0,
            'roles_created': 0,
            'profiles_created': 0,
            'errors': 0
        }
        
        self.logger.info("데이터베이스 마이그레이션 시작")
        
        # 직군/직종 생성
        categories_map = {}
        types_map = {}
        
        with transaction.atomic():
            # 직군 생성
            for cat_name, cat_data in JobProfilePDFParser(None, None).job_categories.items():
                category, created = JobCategory.objects.get_or_create(
                    name=cat_name,
                    defaults={'is_active': True}
                )
                categories_map[cat_data['id']] = category
                if created:
                    stats['categories_created'] += 1
                
                # 직종 생성
                for type_name in cat_data['job_types']:
                    job_type, created = JobType.objects.get_or_create(
                        category=category,
                        name=type_name,
                        defaults={'is_active': True}
                    )
                    type_id = f"{cat_data['id']}_{type_name.replace('/', '_')}"
                    types_map[type_id] = job_type
                    if created:
                        stats['types_created'] += 1
        
        # 직무 프로필 생성
        with tqdm(total=len(jobs_data), desc="DB 마이그레이션") as pbar:
            for job_data in jobs_data:
                try:
                    if not job_data['is_valid']:
                        self.logger.warning(f"검증 실패로 건너뜀: {job_data.get('job_title')}")
                        stats['errors'] += 1
                        pbar.update(1)
                        continue
                    
                    with transaction.atomic():
                        # 직무 역할 생성
                        job_type = types_map.get(job_data.get('type_id'))
                        if not job_type:
                            self.logger.error(f"직종을 찾을 수 없음: {job_data.get('type_id')}")
                            stats['errors'] += 1
                            pbar.update(1)
                            continue
                        
                        job_role, created = JobRole.objects.get_or_create(
                            job_type=job_type,
                            name=job_data['job_title'],
                            defaults={'is_active': True}
                        )
                        if created:
                            stats['roles_created'] += 1
                        
                        # 직무 프로필 생성
                        responsibilities_text = '\n'.join(job_data.get('responsibilities', []))
                        qualifications_text = '\n'.join(job_data.get('qualifications', []))
                        preferred_text = '\n'.join(job_data.get('preferred', []))
                        
                        job_profile, created = JobProfile.objects.update_or_create(
                            job_role=job_role,
                            defaults={
                                'role_responsibility': responsibilities_text,
                                'qualification': qualifications_text,
                                'preferred_qualification': preferred_text,
                                'is_active': True
                            }
                        )
                        
                        if created:
                            stats['profiles_created'] += 1
                            self.logger.info(f"생성됨: {job_data['job_title']}")
                        else:
                            self.logger.info(f"업데이트됨: {job_data['job_title']}")
                
                except Exception as e:
                    self.logger.error(f"마이그레이션 오류 - {job_data.get('job_title')}: {str(e)}")
                    stats['errors'] += 1
                
                pbar.update(1)
        
        self.logger.info(f"마이그레이션 완료: {stats}")
        return stats
    
    def export_to_json(self, jobs_data: List[Dict]) -> str:
        """JSON 파일로 내보내기"""
        json_dir = os.path.join(self.output_dir, 'json')
        os.makedirs(json_dir, exist_ok=True)
        
        # 전체 데이터
        all_jobs_file = os.path.join(json_dir, 'all_job_profiles.json')
        with open(all_jobs_file, 'w', encoding='utf-8') as f:
            json.dump(jobs_data, f, ensure_ascii=False, indent=2)
        
        # 직군별 분리
        by_category = {}
        for job in jobs_data:
            category = job.get('mapped_category', 'undefined')
            if category not in by_category:
                by_category[category] = []
            by_category[category].append(job)
        
        for category, jobs in by_category.items():
            category_file = os.path.join(json_dir, f'{category.replace("/", "_")}_jobs.json')
            with open(category_file, 'w', encoding='utf-8') as f:
                json.dump(jobs, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"JSON 내보내기 완료: {json_dir}")
        return all_jobs_file
    
    def export_to_excel(self, jobs_data: List[Dict]) -> str:
        """Excel 파일로 내보내기"""
        excel_dir = os.path.join(self.output_dir, 'excel')
        os.makedirs(excel_dir, exist_ok=True)
        
        excel_file = os.path.join(excel_dir, 'job_profiles_migration.xlsx')
        
        # 데이터프레임 생성
        df_data = []
        for job in jobs_data:
            df_data.append({
                '페이지': job.get('page_number', ''),
                '직군': job.get('mapped_category', ''),
                '직종': job.get('mapped_type', ''),
                '직무명': job.get('job_title', ''),
                '역할 및 책임': '\n'.join(job.get('responsibilities', [])),
                '자격요건': '\n'.join(job.get('qualifications', [])),
                '우대사항': '\n'.join(job.get('preferred', [])),
                '검증상태': '유효' if job.get('is_valid') else '검증실패',
                '검증오류': ', '.join(job.get('validation_errors', []))
            })
        
        df = pd.DataFrame(df_data)
        
        # Excel Writer
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 전체 데이터 시트
            df.to_excel(writer, sheet_name='전체직무', index=False)
            
            # 직군별 시트
            for category in df['직군'].unique():
                if pd.notna(category):
                    category_df = df[df['직군'] == category]
                    sheet_name = category.replace('/', '_')[:31]  # Excel 시트명 제한
                    category_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 검증 실패 시트
            invalid_df = df[df['검증상태'] == '검증실패']
            if not invalid_df.empty:
                invalid_df.to_excel(writer, sheet_name='검증실패', index=False)
        
        # 스타일 적용
        wb = openpyxl.load_workbook(excel_file)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 헤더 스타일
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 행 높이 조정 (긴 텍스트용)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value and '\n' in str(cell.value):
                        ws.row_dimensions[cell.row].height = 60
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        wb.save(excel_file)
        self.logger.info(f"Excel 내보내기 완료: {excel_file}")
        return excel_file
    
    def generate_migration_report(self, jobs_data: List[Dict], stats: Dict) -> str:
        """마이그레이션 리포트 생성"""
        report_file = os.path.join(self.output_dir, 'migration_report.md')
        
        # 통계 계산
        total_jobs = len(jobs_data)
        valid_jobs = sum(1 for job in jobs_data if job['is_valid'])
        invalid_jobs = total_jobs - valid_jobs
        
        # 직군별 통계
        category_stats = {}
        for job in jobs_data:
            category = job.get('mapped_category', 'undefined')
            if category not in category_stats:
                category_stats[category] = {'total': 0, 'valid': 0}
            category_stats[category]['total'] += 1
            if job['is_valid']:
                category_stats[category]['valid'] += 1
        
        # 리포트 생성
        report_content = f"""# OK금융그룹 직무기술서 마이그레이션 리포트

## 📊 전체 요약
- 생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- PDF 파일: {os.path.basename(self.output_dir.replace('_migrated', '.pdf'))}
- 총 직무 수: {total_jobs}개
- 유효 직무: {valid_jobs}개 ({valid_jobs/total_jobs*100:.1f}%)
- 검증 실패: {invalid_jobs}개

## 📈 데이터베이스 마이그레이션 결과
- 직군 생성: {stats.get('categories_created', 0)}개
- 직종 생성: {stats.get('types_created', 0)}개
- 직무 생성: {stats.get('roles_created', 0)}개
- 프로필 생성: {stats.get('profiles_created', 0)}개
- 오류 발생: {stats.get('errors', 0)}건

## 📁 직군별 현황
"""
        
        for category, stat in sorted(category_stats.items()):
            report_content += f"""
### {category}
- 총 직무: {stat['total']}개
- 유효: {stat['valid']}개 ({stat['valid']/stat['total']*100:.1f}%)
"""
        
        # 검증 실패 상세
        if invalid_jobs > 0:
            report_content += "\n## ⚠️ 검증 실패 직무\n"
            for job in jobs_data:
                if not job['is_valid']:
                    errors = ', '.join(job['validation_errors'])
                    report_content += f"- {job.get('job_title', 'Unknown')}: {errors}\n"
        
        # 출력 파일 정보
        report_content += f"""
## 📤 출력 파일
- JSON: {self.output_dir}/json/all_job_profiles.json
- Excel: {self.output_dir}/excel/job_profiles_migration.xlsx
- 로그: {self.output_dir}/logs/

## ✅ 다음 단계
1. 검증 실패 직무 수동 확인 및 수정
2. 데이터베이스 백업
3. 프로덕션 환경 적용
"""
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report_content)
        
        self.logger.info(f"마이그레이션 리포트 생성: {report_file}")
        return report_file


def main():
    """메인 실행 함수"""
    # 설정
    pdf_path = r"C:/Users/apro/OneDrive/Desktop/설명회자료/OK_Job Profile.pdf"
    output_dir = r"C:/Users/apro/OneDrive/Desktop/설명회자료/job_profile_migrated"
    
    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    # 로깅 설정
    logger = setup_logging(output_dir)
    logger.info("="*60)
    logger.info("OK금융그룹 직무기술서 마이그레이션 시작")
    logger.info("="*60)
    
    try:
        # 1. PDF 파싱
        parser = JobProfilePDFParser(pdf_path, output_dir)
        jobs_data = parser.parse_all_jobs()
        
        if not jobs_data:
            logger.error("파싱된 직무가 없습니다.")
            return
        
        # 2. 데이터베이스 마이그레이션
        migrator = JobProfileMigrator(output_dir)
        
        # Django 환경 확인
        try:
            stats = migrator.migrate_to_database(jobs_data)
        except Exception as e:
            logger.warning(f"데이터베이스 마이그레이션 실패: {str(e)}")
            logger.info("JSON/Excel 내보내기만 진행합니다.")
            stats = {'errors': len(jobs_data)}
        
        # 3. JSON 내보내기
        json_file = migrator.export_to_json(jobs_data)
        
        # 4. Excel 내보내기
        excel_file = migrator.export_to_excel(jobs_data)
        
        # 5. 리포트 생성
        report_file = migrator.generate_migration_report(jobs_data, stats)
        
        # 완료 메시지
        logger.info("="*60)
        logger.info("마이그레이션 완료!")
        logger.info(f"출력 디렉토리: {output_dir}")
        logger.info(f"리포트 확인: {report_file}")
        logger.info("="*60)
        
        # 결과 출력
        print("\n✅ 마이그레이션 완료!")
        print(f"📁 출력 디렉토리: {output_dir}")
        print(f"📄 JSON: {json_file}")
        print(f"📊 Excel: {excel_file}")
        print(f"📋 리포트: {report_file}")
        
    except Exception as e:
        logger.error(f"마이그레이션 중 오류 발생: {str(e)}", exc_info=True)
        raise


if __name__ == '__main__':
    # 필요 패키지 확인
    required_packages = ['PyMuPDF', 'pandas', 'openpyxl', 'tqdm']
    missing_packages = []
    
    for package in required_packages:
        try:
            __import__(package.lower().replace('pymupdf', 'fitz'))
        except ImportError:
            missing_packages.append(package)
    
    if missing_packages:
        print(f"⚠️  필요한 패키지를 설치해주세요:")
        print(f"pip install {' '.join(missing_packages)}")
        sys.exit(1)
    
    main()